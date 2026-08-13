"""
Relatório de Baixa - visão dentro do Atlas de todas as baixas
operacionais (Avaria, Vencimento, Descarte, Degustação, etc.) importadas
do sistema Lovable, de qualquer status (Pendente, Aprovada, Reprovada).
Diferente de baixas_operacionais.py (a lógica de importação/casamento),
este router só lê o que já foi importado - é pra tela de relatório, não
pra receber webhook."""
import io
from collections import defaultdict
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Body, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font
import pandas as pd

from .. import models
from ..database import get_db
from ..deps import obter_usuario_atual, requer_papel
from ..baixas_operacionais import sincronizar_com_lovable, SincronizacaoIndisponivel, importar_lote, importar_planilha_historico_lovable

router = APIRouter(prefix="/baixas-operacionais", tags=["baixas_operacionais"])


@router.post("/resetar")
def resetar_historico(
    usuario: models.Usuario = Depends(requer_papel("admin")),
    db: Session = Depends(get_db),
):
    """Apaga TODO o histórico de baixas operacionais guardado no Atlas
    (tabela baixas_operacionais) - usado quando o relatório fica com uma
    contagem que não bate (ex: linhas de uma sincronização/importação
    antiga, de antes de alguma correção de dedup, que nunca foram
    corrigidas ou removidas). Em vez de tentar consertar linha a linha,
    apaga tudo daqui e deixa o Lovable (fonte da verdade) repovoar do
    zero na próxima chamada a /sincronizar. Não toca em Divergência nem em
    nenhuma outra tabela - só nas baixas em si. Restrito a admin porque é
    uma ação destrutiva sem confirmação extra na tela."""
    total_apagado = db.query(models.BaixaOperacional).delete()
    db.commit()
    return {"apagadas": total_apagado}


@router.post("/importar-lote")
def importar_lote_colado(
    payload: dict = Body(...),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Fecha manualmente uma lacuna histórica entre o que existe no
    Lovable e o que já foi importado pro Atlas (ex: linhas de antes do
    webhook automático existir, ou alguma falha pontual de entrega).
    Protegido por login (mesmo usuário logado na tela, não pela chave de
    integração - por isso fica aqui, junto do relatório, e não em
    integracoes_router.py) porque quem aciona isso é uma pessoa colando
    um export tirado na mão do SQL editor do Lovable, não um sistema
    automático. Espera {"registros": [ {...}, ... ]} - mesmo formato de
    linha que .../integracoes/lovable/baixas/lote. Upsert por origem_id,
    então pode ser rodado de novo com o mesmo lote sem duplicar nada."""
    registros = payload.get("registros")
    if not isinstance(registros, list):
        raise HTTPException(400, "Payload precisa ter uma lista em 'registros'.")
    resultado = importar_lote(db, registros)
    db.commit()
    return resultado


@router.post("/sincronizar")
def sincronizar(
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Botão "Sincronizar agora" da tela Relatório de Baixa: busca ao vivo
    o estado atual da tabela baixa_operacional no Supabase do Lovable e
    reimporta tudo pro Atlas (upsert por origem_id - atualiza o que mudou
    de status lá, ex: Pendente -> Aprovada, sem duplicar nada)."""
    try:
        resultado = sincronizar_com_lovable(db)
    except SincronizacaoIndisponivel as e:
        raise HTTPException(500, str(e))
    db.commit()
    return resultado


@router.post("/reconciliar-planilha")
async def reconciliar_planilha_historico(
    arquivo: UploadFile = File(...),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Fecha a lacuna entre o que já está no Atlas e o export "Baixar
    relatório completo" da tela Baixas Operacionais (Lovable) - pra usar
    quando a sincronização automática (webhook e/ou "Sincronizar agora")
    não trouxe tudo, ex: LOVABLE_SYNC_URL/LOVABLE_SYNC_SECRET nunca
    configuradas no Render. Ver importar_planilha_historico_lovable (baixas_operacionais.py)
    pro porquê disso ser reconciliação por assinatura, não upsert por id -
    a planilha não carrega o uuid único da linha no Lovable."""
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")

    aba = "Histórico" if "Histórico" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[aba]
    linhas_brutas = list(ws.iter_rows(values_only=True))
    if not linhas_brutas:
        raise HTTPException(400, "Planilha vazia.")

    cabecalho = [str(c).strip() if c else "" for c in linhas_brutas[0]]
    esperado = ["Data", "Código", "Almoxarifado", "Quantidade", "Valor Total", "Motivo", "Status"]
    faltando = [c for c in esperado if c not in cabecalho]
    if faltando:
        raise HTTPException(400, f"Colunas esperadas não encontradas: {faltando}. Cabeçalho: {cabecalho}")

    linhas = [dict(zip(cabecalho, linha)) for linha in linhas_brutas[1:]]
    resultado = importar_planilha_historico_lovable(db, linhas)
    db.commit()
    return resultado


@router.get("")
def listar_baixas(
    status_fluxo: str | None = Query(None, description="PENDENTE | APROVADA | REPROVADA"),
    almoxarifado: str | None = None,
    hipotese_aplicada: str | None = None,
    data_inicio: date | None = None,
    data_fim: date | None = None,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    q = db.query(models.BaixaOperacional)
    if status_fluxo:
        q = q.filter(models.BaixaOperacional.status_fluxo == status_fluxo.upper())
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if hipotese_aplicada:
        q = q.filter(models.BaixaOperacional.hipotese_aplicada == hipotese_aplicada)
    if data_inicio:
        q = q.filter(models.BaixaOperacional.data_baixa >= data_inicio)
    if data_fim:
        q = q.filter(models.BaixaOperacional.data_baixa <= data_fim)

    linhas = q.order_by(models.BaixaOperacional.data_baixa.desc()).all()

    resumo = {
        "total": len(linhas),
        "pendentes": sum(1 for l in linhas if l.status_fluxo == "PENDENTE"),
        "aprovadas": sum(1 for l in linhas if l.status_fluxo == "APROVADA"),
        "reprovadas": sum(1 for l in linhas if l.status_fluxo == "REPROVADA"),
        "resolvidas_automaticamente": sum(1 for l in linhas if l.divergencia_vinculada_id),
        "aguardando_divergencia": sum(1 for l in linhas if l.status_fluxo == "APROVADA" and not l.divergencia_vinculada_id),
        "valor_total": sum(l.valor_total or 0 for l in linhas),
    }

    itens = [
        {
            "id": l.id,
            "sku": l.sku,
            "almoxarifado": l.almoxarifado,
            "almoxarifado_origem": l.almoxarifado_origem,
            "motivo": l.motivo_baixa_bruto,
            "hipotese_aplicada": l.hipotese_aplicada,
            "quantidade": l.quantidade,
            "valor_total": l.valor_total,
            "status_fluxo": l.status_fluxo,
            "solicitante_nome": l.solicitante_nome,
            "data_baixa": l.data_baixa,
            "divergencia_vinculada_id": l.divergencia_vinculada_id,
            "recebido_em": l.recebido_em,
        }
        for l in linhas
    ]
    return {"resumo": resumo, "itens": itens}


# ---------------------------------------------------------------------------
# Dashboard "Mapeamento de Passivos" - visão gerencial de todas as baixas
# operacionais já trazidas do Lovable (não é a origem/webhook, é análise
# sobre o que já está importado). Cobre:
#   1) status operacionais usados (Pendente/Aprovada/Reprovada) - pizza;
#   2) de onde veio o mapeamento de cada baixa aprovada: de uma
#      divergência achada no Inventário Mensal (Divergencia.origem ==
#      "fechamento_inventario"), de uma divergência do dia a dia
#      (Divergencia.origem == "movimentacao"), aprovada mas ainda sem
#      casar com nenhuma divergência, ou nem chegou a ser decidida
#      (Pendente/Reprovada) - coluna;
#   3) evolução mês a mês (MoM) do valor de baixas aplicadas + taxa de
#      resolução automática - coluna+linha;
#   4) Top 10 SKUs mais recorrentes em baixa e Top 10 por impacto
#      financeiro - tabelas.
# Todo clique num gráfico/linha/linha de tabela abre o mesmo pop de
# drill-down (GET .../dashboard/itens) filtrado pela categoria clicada.
# ---------------------------------------------------------------------------

CATEGORIA_MAPEAMENTO_LABELS = {
    "inventario_mensal": "Mapeada via Inventário Mensal",
    "movimentacao_diaria": "Mapeada via Movimentação Diária",
    "aguardando_divergencia": "Aprovada, aguardando divergência",
    "nao_decidida": "Pendente/Reprovada (não decidida)",
}


def _mapa_divergencias_das_baixas(db: Session, baixas: list) -> dict:
    ids = {b.divergencia_vinculada_id for b in baixas if b.divergencia_vinculada_id}
    if not ids:
        return {}
    return {d.id: d for d in db.query(models.Divergencia).filter(models.Divergencia.id.in_(ids)).all()}


def _categoria_mapeamento(baixa, divergencias_por_id: dict) -> str:
    if not baixa.divergencia_vinculada_id:
        return "aguardando_divergencia" if baixa.status_fluxo == "APROVADA" else "nao_decidida"
    div = divergencias_por_id.get(baixa.divergencia_vinculada_id)
    if div and div.origem == "fechamento_inventario":
        return "inventario_mensal"
    return "movimentacao_diaria"


def _filtrar_baixas_dashboard(
    db: Session, status_fluxo: str | None, categoria_mapeamento: str | None, mes: str | None, sku: str | None,
    motivo: str | None = None, excluir_categoria_mapeamento: str | None = None,
    ano: int | None = None, data_inicio=None, data_fim=None, almoxarifado: str | None = None,
    mes_numero: int | None = None,
) -> list:
    q = db.query(models.BaixaOperacional)
    if status_fluxo:
        q = q.filter(models.BaixaOperacional.status_fluxo == status_fluxo.upper())
    if sku:
        q = q.filter(models.BaixaOperacional.sku == sku)
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivo:
        # aceita uma lista separada por vírgula (clique no segmento "Outros" do
        # gráfico de motivos, que agrupa vários motivos de menor valor)
        motivos_lista = [m.strip() for m in motivo.split(",") if m.strip()]
        q = q.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    baixas = q.all()
    if mes:
        baixas = [b for b in baixas if b.data_baixa and str(b.data_baixa)[:7] == mes]
    if ano or mes_numero or data_inicio or data_fim:
        # filtro de recorte do painel (Data/Mês/Ano) - o `mes` acima é o mês exato
        # (YYYY-MM) de um clique num gráfico, um conceito diferente/independente do
        # `mes_numero` (1-12, qualquer ano) do filtro do painel.
        baixas = [b for b in baixas if _data_no_periodo(b.data_baixa, ano, mes_numero, data_inicio, data_fim)]
    if categoria_mapeamento or excluir_categoria_mapeamento:
        divergencias_por_id = _mapa_divergencias_das_baixas(db, baixas)
        if categoria_mapeamento:
            baixas = [b for b in baixas if _categoria_mapeamento(b, divergencias_por_id) == categoria_mapeamento]
        if excluir_categoria_mapeamento:
            baixas = [b for b in baixas if _categoria_mapeamento(b, divergencias_por_id) != excluir_categoria_mapeamento]
    return baixas


@router.get("/dashboard/kpis")
def dashboard_passivos_kpis(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    baixas = db.query(models.BaixaOperacional).all()
    divergencias_por_id = _mapa_divergencias_das_baixas(db, baixas)

    por_categoria = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in baixas:
        cat = _categoria_mapeamento(b, divergencias_por_id)
        por_categoria[cat]["quantidade"] += 1
        por_categoria[cat]["valor"] += b.valor_total or 0

    aprovadas = [b for b in baixas if b.status_fluxo == "APROVADA"]
    mapeamento = {
        chave: {"label": rotulo, "quantidade": por_categoria[chave]["quantidade"], "valor": round(por_categoria[chave]["valor"], 2)}
        for chave, rotulo in CATEGORIA_MAPEAMENTO_LABELS.items()
    }

    return {
        "total": len(baixas),
        "valor_total": round(sum(b.valor_total or 0 for b in baixas), 2),
        "pendentes": sum(1 for b in baixas if b.status_fluxo == "PENDENTE"),
        "aprovadas": len(aprovadas),
        "reprovadas": sum(1 for b in baixas if b.status_fluxo == "REPROVADA"),
        "valor_aprovadas": round(sum(b.valor_total or 0 for b in aprovadas), 2),
        "mapeamento": mapeamento,
    }


@router.get("/dashboard/status-pizza")
def dashboard_passivos_status_pizza(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Status operacionais usados (Pendente/Aprovada/Reprovada) - todas as
    baixas já trazidas do Lovable, de qualquer origem de mapeamento."""
    baixas = db.query(models.BaixaOperacional).all()
    por_status = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in baixas:
        chave = b.status_fluxo or "NAO_INFORMADO"
        por_status[chave]["quantidade"] += 1
        por_status[chave]["valor"] += b.valor_total or 0
    rotulos = {"PENDENTE": "Pendente", "APROVADA": "Aprovada", "REPROVADA": "Reprovada", "NAO_INFORMADO": "Não informado"}
    return [
        {"status_fluxo": k, "label": rotulos.get(k, k), "quantidade": v["quantidade"], "valor": round(v["valor"], 2)}
        for k, v in por_status.items()
    ]


TOP_N_MOTIVOS = 6  # motivos individuais mostrados no gráfico; o resto agrupa em "Outros"


@router.get("/dashboard/motivos-mensal")
def dashboard_passivos_motivos_mensal(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db),
):
    """Principais motivos de baixa (Avaria, Vencimento, Descarte,
    Degustação, Perda/Furto etc.) mês a mês, por valor - substitui a pizza
    de "Status Operacionais Usados". Só baixas APROVADAS (as que de fato
    foram aplicadas) e SEM contar o que já é ajuste de inventário mensal
    (categoria_mapeamento == "inventario_mensal" - esse fluxo já tem seu
    próprio painel dedicado, "Fluxo de Inventário", não faz sentido
    duplicar aqui). Os TOP_N_MOTIVOS de maior valor total aparecem
    individualmente; o resto entra agrupado em "Outros" pra não estourar a
    legenda do gráfico. Respeita o mesmo recorte (Data/Mês/Ano/
    Almoxarifado/Motivo) do Resumo Executivo (13/08/2026) - antes só os 2
    cards de KPI respondiam ao filtro, os gráficos ficavam sempre com a
    base inteira."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    motivos_lista = _parse_motivos(motivo)

    q = db.query(models.BaixaOperacional).filter(models.BaixaOperacional.status_fluxo == "APROVADA")
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivos_lista:
        q = q.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    aprovadas = [b for b in q.all() if _data_no_periodo(b.data_baixa, ano, mes, di, df)]
    divergencias_por_id = _mapa_divergencias_das_baixas(db, aprovadas)
    aprovadas = [b for b in aprovadas if _categoria_mapeamento(b, divergencias_por_id) != "inventario_mensal"]

    valor_por_motivo_total = defaultdict(float)
    por_mes_motivo = defaultdict(lambda: defaultdict(float))
    meses = set()
    for b in aprovadas:
        if not b.data_baixa:
            continue
        motivo = b.motivo_baixa_bruto or "Não informado"
        mes = str(b.data_baixa)[:7]
        meses.add(mes)
        valor_por_motivo_total[motivo] += b.valor_total or 0
        por_mes_motivo[mes][motivo] += b.valor_total or 0

    motivos_ordenados = sorted(valor_por_motivo_total.keys(), key=lambda m: valor_por_motivo_total[m], reverse=True)
    motivos_top = motivos_ordenados[:TOP_N_MOTIVOS]
    tem_outros = len(motivos_ordenados) > TOP_N_MOTIVOS
    motivos_finais = motivos_top + (["Outros"] if tem_outros else [])

    meses_ordenados = sorted(meses)
    valores = {motivo: [] for motivo in motivos_finais}
    for mes in meses_ordenados:
        valores_mes_outros = 0.0
        for motivo in motivos_ordenados:
            v = round(por_mes_motivo[mes].get(motivo, 0.0), 2)
            if motivo in motivos_top:
                valores[motivo].append(v)
            else:
                valores_mes_outros += v
        if tem_outros:
            valores["Outros"].append(round(valores_mes_outros, 2))

    return {
        "meses": meses_ordenados,
        "motivos": motivos_finais,
        "valores": valores,
        "motivos_agrupados_em_outros": motivos_ordenados[TOP_N_MOTIVOS:] if tem_outros else [],
    }


@router.get("/dashboard/motivos-resumo")
def dashboard_passivos_motivos_resumo(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db),
):
    """Resumo por motivo de baixa (Avaria, Vencimento, Descarte...) com
    quantidade e custo total - TODOS os motivos individualmente, sem o
    top-N/"Outros" do gráfico de motivos-mensal (aqui é uma tabela, não tem
    o limite de espaço de uma legenda). Mesmo filtro do gráfico: só
    aprovadas, sem contar o que já é ajuste de inventário mensal (esse
    fluxo já tem painel dedicado - Fluxo de Inventário). Respeita o mesmo
    recorte do Resumo Executivo (13/08/2026, ver motivos-mensal acima)."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    motivos_lista = _parse_motivos(motivo)

    q = db.query(models.BaixaOperacional).filter(models.BaixaOperacional.status_fluxo == "APROVADA")
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivos_lista:
        q = q.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    aprovadas = [b for b in q.all() if _data_no_periodo(b.data_baixa, ano, mes, di, df)]
    divergencias_por_id = _mapa_divergencias_das_baixas(db, aprovadas)
    aprovadas = [b for b in aprovadas if _categoria_mapeamento(b, divergencias_por_id) != "inventario_mensal"]

    por_motivo = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in aprovadas:
        motivo = b.motivo_baixa_bruto or "Não informado"
        por_motivo[motivo]["quantidade"] += 1
        por_motivo[motivo]["valor"] += b.valor_total or 0

    resumo = [
        {"motivo": motivo, "quantidade": v["quantidade"], "valor": round(v["valor"], 2)}
        for motivo, v in por_motivo.items()
    ]
    resumo.sort(key=lambda x: abs(x["valor"]), reverse=True)
    return resumo


@router.get("/dashboard/mapeamento-origem")
def dashboard_passivos_mapeamento_origem(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """De onde vem o mapeamento de cada baixa - Inventário Mensal vs
    Movimentação Diária vs ainda sem decisão/casamento. Ver
    CATEGORIA_MAPEAMENTO_LABELS e _categoria_mapeamento acima."""
    baixas = db.query(models.BaixaOperacional).all()
    divergencias_por_id = _mapa_divergencias_das_baixas(db, baixas)
    por_categoria = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in baixas:
        cat = _categoria_mapeamento(b, divergencias_por_id)
        por_categoria[cat]["quantidade"] += 1
        por_categoria[cat]["valor"] += b.valor_total or 0
    return [
        {"categoria": chave, "label": rotulo, "quantidade": por_categoria[chave]["quantidade"], "valor": round(por_categoria[chave]["valor"], 2)}
        for chave, rotulo in CATEGORIA_MAPEAMENTO_LABELS.items()
    ]


@router.get("/dashboard/evolucao-mensal")
def dashboard_passivos_evolucao_mensal(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db),
):
    """MoM de todas as baixas APROVADAS (as que de fato foram aplicadas -
    Pendente/Reprovada não é passivo real ainda/nunca vai ser), mês a mês
    por data_baixa: valor total (coluna) e taxa de resolução automática -
    % que casou com uma divergência, seja de inventário mensal seja de
    movimentação (linha) - a "curva de evolução do processo" pedida.
    Também traz, por mês, o resultado do fluxo de inventário (Entradas -
    Saídas de TODOS os inventários, ver _fluxo_inventario_por_mes) somado
    ao lado do mapeamento de passivos, pra ficar tudo na mesma visão MoM.
    Respeita o mesmo recorte (Data/Mês/Ano/Almoxarifado/Motivo) do Resumo
    Executivo (13/08/2026) - o `mes` aqui é o filtro do painel (1-12), não
    um "YYYY-MM" exato."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    motivos_lista = _parse_motivos(motivo)

    qb = db.query(models.BaixaOperacional).filter(models.BaixaOperacional.status_fluxo == "APROVADA")
    if almoxarifado:
        qb = qb.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivos_lista:
        qb = qb.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    aprovadas = [b for b in qb.all() if _data_no_periodo(b.data_baixa, ano, mes, di, df)]
    por_mes = defaultdict(lambda: {"quantidade": 0, "valor": 0.0, "resolvidas": 0})
    for b in aprovadas:
        if not b.data_baixa:
            continue
        mes_baixa = str(b.data_baixa)[:7]
        por_mes[mes_baixa]["quantidade"] += 1
        por_mes[mes_baixa]["valor"] += b.valor_total or 0
        if b.divergencia_vinculada_id:
            por_mes[mes_baixa]["resolvidas"] += 1

    fluxo_por_mes = _fluxo_inventario_por_mes(db, ano=ano, mes_numero=mes, data_inicio=di, data_fim=df, almoxarifado=almoxarifado)
    todos_meses = sorted(set(por_mes.keys()) | set(fluxo_por_mes.keys()))

    resultado = []
    for mes_str in todos_meses:
        v = por_mes[mes_str]
        taxa = round(v["resolvidas"] / v["quantidade"] * 100, 1) if v["quantidade"] else None
        fluxo = fluxo_por_mes.get(mes_str, {"entradas_valor": 0.0, "saidas_valor": 0.0, "resultado_valor": 0.0})
        resultado.append({
            "mes": mes_str, "quantidade": v["quantidade"], "valor": round(v["valor"], 2), "taxa_resolucao_automatica_pct": taxa,
            "resultado_inventario_mes": round(fluxo["resultado_valor"], 2),
            "entradas_inventario_mes": round(fluxo["entradas_valor"], 2),
            "saidas_inventario_mes": round(fluxo["saidas_valor"], 2),
        })
    return resultado


# ---------------------------------------------------------------------------
# Fluxo de Inventário (Entradas x Saídas x Resultado) - "Mapeamento de grana
# de TODOS os inventários": Total de Entradas - Total de Saídas do mês =
# Resultado Total do Mês. Base: AjusteInventarioOficial (a tabela OFICIAL de
# ajustes já conciliados - "Ace4"/aba "Estoque" da planilha Inventários),
# NÃO o fechamento bruto (ItemFechamento) - esse incluía item que a
# operação sinalizou como divergente mas nunca ajustou de fato, por
# problema de processo. Só entra no cálculo quem tem
# conta_como_ajuste_inventario=True (ver ajustes_inventario_router.py: "Sim"
# explícito na planilha, ou qualquer lançamento a partir de jul/2026,
# quando baixa de passivo passou a vir só por nota fiscal). ajuste_qtd > 0
# = sobra encontrada = Entrada; ajuste_qtd < 0 = falta = Saída.
# ---------------------------------------------------------------------------

def _direcao_ajuste(registro) -> str | None:
    if not registro.ajuste_qtd:
        return None
    return "entrada" if registro.ajuste_qtd > 0 else "saida"


def _ajustes_inventario_contados(
    db: Session, mes: str | None = None, direcao: str | None = None,
    ano: int | None = None, mes_numero: int | None = None, data_inicio=None, data_fim=None,
    almoxarifado: str | None = None,
) -> list:
    """[(AjusteInventarioOficial, mes, direcao), ...] só dos registros que
    contam como ajuste de inventário (ver conta_como_ajuste_inventario).
    `mes` (YYYY-MM exato) é o filtro do drill-down de um clique específico -
    `ano`/`mes_numero`/`data_inicio`/`data_fim`/`almoxarifado` são o recorte
    do painel (Resumo Executivo), independentes um do outro (13/08/2026)."""
    q = db.query(models.AjusteInventarioOficial).filter(models.AjusteInventarioOficial.conta_como_ajuste_inventario.is_(True))
    if almoxarifado:
        q = q.filter(models.AjusteInventarioOficial.almoxarifado == almoxarifado)
    resultado = []
    for registro in q.all():
        if not registro.dt_invent or not registro.ajuste_qtd:
            continue
        if not _data_no_periodo(registro.dt_invent, ano, mes_numero, data_inicio, data_fim):
            continue
        mes_registro = str(registro.dt_invent)[:7]
        if mes and mes_registro != mes:
            continue
        direcao_registro = _direcao_ajuste(registro)
        if not direcao_registro:
            continue
        if direcao and direcao_registro != direcao:
            continue
        resultado.append((registro, mes_registro, direcao_registro))
    return resultado


def _fluxo_inventario_por_mes(
    db: Session, ano: int | None = None, mes_numero: int | None = None, data_inicio=None, data_fim=None,
    almoxarifado: str | None = None,
) -> dict:
    """{mes: {entradas_valor, saidas_valor, resultado_valor, entradas_qtd, saidas_qtd}}
    somando os inventários do recorte pedido (sem filtro = todos os
    almoxarifados/período, como sempre foi)."""
    linhas = _ajustes_inventario_contados(db, ano=ano, mes_numero=mes_numero, data_inicio=data_inicio, data_fim=data_fim, almoxarifado=almoxarifado)
    por_mes = defaultdict(lambda: {"entradas_valor": 0.0, "saidas_valor": 0.0, "entradas_qtd": 0, "saidas_qtd": 0})
    for registro, mes_registro, direcao_registro in linhas:
        valor = abs(registro.valor_total or 0)
        if direcao_registro == "entrada":
            por_mes[mes_registro]["entradas_valor"] += valor
            por_mes[mes_registro]["entradas_qtd"] += 1
        else:
            por_mes[mes_registro]["saidas_valor"] += valor
            por_mes[mes_registro]["saidas_qtd"] += 1
    for v in por_mes.values():
        v["resultado_valor"] = v["entradas_valor"] - v["saidas_valor"]
    return dict(por_mes)


@router.get("/dashboard/fluxo-inventario-mensal")
def dashboard_fluxo_inventario_mensal(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db),
):
    """Total de Entradas - Total de Saídas do mês = Resultado Total do Mês,
    mês a mês, respeitando o recorte (Data/Mês/Ano/Almoxarifado) do Resumo
    Executivo (13/08/2026) - sem filtro, soma todos os almoxarifados/
    período, como sempre foi. Motivo não se aplica aqui (não existe na
    tabela de ajustes de inventário)."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    fluxo_por_mes = _fluxo_inventario_por_mes(db, ano=ano, mes_numero=mes, data_inicio=di, data_fim=df, almoxarifado=almoxarifado)
    resultado = []
    for mes_str in sorted(fluxo_por_mes.keys()):
        v = fluxo_por_mes[mes_str]
        resultado.append({
            "mes": mes_str,
            "entradas_valor": round(v["entradas_valor"], 2), "saidas_valor": round(v["saidas_valor"], 2),
            "resultado_valor": round(v["entradas_valor"] - v["saidas_valor"], 2),
            "entradas_qtd": v["entradas_qtd"], "saidas_qtd": v["saidas_qtd"],
        })
    return resultado


@router.get("/dashboard/fluxo-inventario-totais")
def dashboard_fluxo_inventario_totais(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """KPIs agregados de TODOS os meses/inventários: Entradas, Saídas e
    Resultado Total - "Mapeamento de grana de todos os inventários"."""
    fluxo_por_mes = _fluxo_inventario_por_mes(db)
    entradas = round(sum(v["entradas_valor"] for v in fluxo_por_mes.values()), 2)
    saidas = round(sum(v["saidas_valor"] for v in fluxo_por_mes.values()), 2)
    return {
        "entradas_valor": entradas, "saidas_valor": saidas, "resultado_valor": round(entradas - saidas, 2),
        "entradas_qtd": sum(v["entradas_qtd"] for v in fluxo_por_mes.values()),
        "saidas_qtd": sum(v["saidas_qtd"] for v in fluxo_por_mes.values()),
        "meses_com_fechamento": len(fluxo_por_mes),
    }


@router.get("/dashboard/itens-fluxo-inventario")
def dashboard_itens_fluxo_inventario(
    mes: str | None = Query(None, description="YYYY-MM"),
    direcao: str | None = Query(None, description="entrada | saida"),
    ano: int | None = Query(None, description="filtro do painel (recorte atual)"),
    mes_numero: int | None = Query(None, ge=1, le=12, description="filtro do painel (recorte atual, 1-12, qualquer ano) - diferente do `mes` acima"),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD - filtro do painel (recorte atual)"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD - filtro do painel (recorte atual)"),
    almoxarifado: str | None = Query(None, description="filtro do painel (recorte atual)"),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Drill-down do painel de Fluxo de Inventário e da série 'Resultado do
    Fluxo de Inventário' no MoM - lista os ajustes oficiais (SKU x
    almoxarifado x lote) que geraram entrada/saída no mês/direção
    clicados. ano/mes_numero/data_inicio/data_fim/almoxarifado (opcionais)
    deixam o drill-down respeitar também o recorte atual do painel - usado,
    por exemplo, pelo botão "Ver todos os Ajustes de Inventário"."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    linhas = _ajustes_inventario_contados(db, mes=mes, direcao=direcao, ano=ano, mes_numero=mes_numero, data_inicio=di, data_fim=df, almoxarifado=almoxarifado)
    skus = {registro.sku for registro, _m, _d in linhas if registro.sku}
    descricoes = {p.sku: p.descricao for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus)).all()}

    itens = [
        {
            "id": registro.id, "sku": registro.sku, "descricao_produto": registro.descricao_produto or descricoes.get(registro.sku),
            "almoxarifado": registro.almoxarifado, "categoria_produto": registro.categoria_produto,
            "id_lote": registro.id_lote, "id_invent": registro.id_invent,
            "qtd_sistema": registro.qtd_sistema, "qtd_contagem": registro.qtd_contagem, "divergencia_qtd": registro.ajuste_qtd,
            "valor_estimado": abs(registro.valor_total or 0), "direcao": direcao_registro,
            "data_fechamento": str(registro.dt_invent) if registro.dt_invent else None,
        }
        for registro, _mes_registro, direcao_registro in linhas
    ]
    itens.sort(key=lambda x: abs(x["valor_estimado"] or 0), reverse=True)
    return {
        "itens": itens, "total": len(itens),
        "entradas_valor": round(sum(i["valor_estimado"] for i in itens if i["direcao"] == "entrada"), 2),
        "saidas_valor": round(sum(i["valor_estimado"] for i in itens if i["direcao"] == "saida"), 2),
    }


@router.get("/dashboard/classificacao-oficial-inventario")
def dashboard_classificacao_oficial_inventario(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """"Mapeada de passivos" da tabela oficial de ajustes ("Ace4"): de tudo
    que passou pelo módulo de inventário, quanto é ajuste de inventário DE
    FATO (conta_como_ajuste_inventario=True - coluna "Sim" ou lançamento
    pós jul/2026) x quanto é baixa de passivo que só passou por ali e já
    está mapeada em outro lugar (coluna "Não", antes de jul/2026) x quanto
    é legado de antes da separação Sim/Não existir (desconsiderado dos
    dois lados, por decisão do Maurício - não tem como saber se era
    ajuste ou passivo)."""
    todos = db.query(models.AjusteInventarioOficial).all()
    grupos = {
        "ajuste_inventario": {"label": "Ajuste de Inventário (Sim / pós-jul-2026)", "quantidade": 0, "valor": 0.0},
        "passivo_ja_mapeado": {"label": "Baixa de Passivo (Não) - já mapeada em Baixas", "quantidade": 0, "valor": 0.0},
        "legado_desconsiderado": {"label": "Legado pré-separação (desconsiderado)", "quantidade": 0, "valor": 0.0},
    }
    for r in todos:
        flag_lower = str(r.inventario_flag_bruto).strip().lower() if r.inventario_flag_bruto is not None else ""
        if r.conta_como_ajuste_inventario:
            chave = "ajuste_inventario"
        elif flag_lower == "não":
            chave = "passivo_ja_mapeado"
        else:
            chave = "legado_desconsiderado"
        grupos[chave]["quantidade"] += 1
        grupos[chave]["valor"] += r.valor_total or 0
    for g in grupos.values():
        g["valor"] = round(g["valor"], 2)
    return grupos


def _baixas_aprovadas_filtradas(
    db: Session, ano: int | None, mes: int | None, data_inicio, data_fim,
    almoxarifado: str | None, motivos_lista: list[str] | None,
) -> list:
    """Baixas APROVADAS dentro do recorte do painel (Resumo Executivo) -
    reaproveitado pelos Top 10 (13/08/2026, antes ficavam sempre com a base
    inteira, sem responder ao filtro)."""
    q = db.query(models.BaixaOperacional).filter(models.BaixaOperacional.status_fluxo == "APROVADA")
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivos_lista:
        q = q.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    return [b for b in q.all() if _data_no_periodo(b.data_baixa, ano, mes, data_inicio, data_fim)]


def _top_10_por_sku(baixas: list, db: Session, chave_ordenacao) -> list:
    por_sku = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in baixas:
        if not b.sku:
            continue
        por_sku[b.sku]["quantidade"] += 1
        por_sku[b.sku]["valor"] += b.valor_total or 0
    descricoes = {p.sku: p.descricao for p in db.query(models.Produto).filter(models.Produto.sku.in_(por_sku.keys())).all()}
    itens = [{"sku": sku, "descricao_produto": descricoes.get(sku), "quantidade": v["quantidade"], "valor": round(v["valor"], 2)} for sku, v in por_sku.items()]
    itens.sort(key=chave_ordenacao, reverse=True)
    return itens[:10]


@router.get("/dashboard/top-recorrentes")
def dashboard_passivos_top_recorrentes(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db),
):
    """Top 10 SKUs que mais vezes tiveram baixa aprovada - recorrência
    (quantidade de ocorrências), não valor - ver top-impacto-financeiro
    para o ranking por valor. Respeita o recorte do Resumo Executivo
    (13/08/2026)."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    aprovadas = _baixas_aprovadas_filtradas(db, ano, mes, di, df, almoxarifado, _parse_motivos(motivo))
    return _top_10_por_sku(aprovadas, db, chave_ordenacao=lambda x: x["quantidade"])


@router.get("/dashboard/top-impacto-financeiro")
def dashboard_passivos_top_impacto_financeiro(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db),
):
    """Top 10 SKUs com maior valor total baixado (aprovado). Respeita o
    recorte do Resumo Executivo (13/08/2026)."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    aprovadas = _baixas_aprovadas_filtradas(db, ano, mes, di, df, almoxarifado, _parse_motivos(motivo))
    return _top_10_por_sku(aprovadas, db, chave_ordenacao=lambda x: x["valor"])


@router.get("/dashboard/itens")
def dashboard_passivos_itens(
    status_fluxo: str | None = None,
    categoria_mapeamento: str | None = Query(None, description="inventario_mensal | movimentacao_diaria | aguardando_divergencia | nao_decidida"),
    mes: str | None = Query(None, description="YYYY-MM"),
    sku: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula (usado pelo segmento 'Outros' de dashboard/motivos-mensal)"),
    excluir_categoria_mapeamento: str | None = Query(None, description="usado pelo gráfico de motivos, que já exclui inventario_mensal"),
    ano: int | None = Query(None, description="filtro do painel (recorte atual) - diferente do `mes` acima, que é o mês exato clicado"),
    mes_numero: int | None = Query(None, ge=1, le=12, description="filtro do painel (recorte atual, 1-12, qualquer ano) - diferente do `mes` acima"),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD - filtro do painel (recorte atual)"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD - filtro do painel (recorte atual)"),
    almoxarifado: str | None = Query(None, description="filtro do painel (recorte atual)"),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Drill-down usado por TODOS os gráficos/tabelas do dashboard
    Mapeamento de Passivos - clicar numa fatia da pizza, numa coluna do
    mapeamento por origem, num mês do MoM, numa linha de um Top 10 ou num
    segmento do gráfico de motivos de baixa cai aqui, com o filtro
    correspondente. ano/mes_numero/data_inicio/data_fim/almoxarifado
    (opcionais) deixam o drill-down respeitar também o recorte atual do
    painel (Data/Mês/Ano/Almoxarifado), não só o filtro específico do ponto
    clicado - usado, por exemplo, pelo botão "Ver todos os Passivos"."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    baixas = _filtrar_baixas_dashboard(
        db, status_fluxo, categoria_mapeamento, mes, sku, motivo, excluir_categoria_mapeamento,
        ano=ano, data_inicio=di, data_fim=df, almoxarifado=almoxarifado, mes_numero=mes_numero,
    )
    divergencias_por_id = _mapa_divergencias_das_baixas(db, baixas)
    descricoes = {p.sku: p.descricao for p in db.query(models.Produto).filter(models.Produto.sku.in_({b.sku for b in baixas if b.sku})).all()}

    itens = [
        {
            "id": b.id, "sku": b.sku, "descricao_produto": descricoes.get(b.sku),
            "almoxarifado": b.almoxarifado, "motivo": b.motivo_baixa_bruto,
            "quantidade": b.quantidade, "valor_total": b.valor_total,
            "status_fluxo": b.status_fluxo, "solicitante_nome": b.solicitante_nome,
            "data_baixa": str(b.data_baixa) if b.data_baixa else None,
            "divergencia_vinculada_id": b.divergencia_vinculada_id,
            "categoria_mapeamento": _categoria_mapeamento(b, divergencias_por_id),
        }
        for b in baixas
    ]
    itens.sort(key=lambda x: (x["data_baixa"] or "", abs(x["valor_total"] or 0)), reverse=True)
    return {"itens": itens, "total": len(itens), "valor_total": round(sum(b.valor_total or 0 for b in baixas), 2)}


# ---------------------------------------------------------------------------
# Resumo Executivo (12/08/2026) - a tela Mapeamento de Passivos foi
# simplificada pra só 2 indicadores centrais (Passivos e Resultado de
# Inventário Acumulado), com filtro de Data/Mês/Ano, Almoxarifado e Motivo,
# e um pop-up de duplo-clique com o resumo narrado dos dois juntos. Ver
# DECISOES.md / claude/indicadores-passivos.md (projeto) pro contexto dessa
# mudança - substitui os 9 cards + 2 gráficos antigos.
#
# "Passivos" = valor total das baixas operacionais APROVADAS no recorte
# filtrado (mesmo número que já existia em VALOR_TOTAL_APROVADAS, só que
# agora filtrável). "Resultado de Inventário Acumulado" = Entradas − Saídas
# de TODOS os ajustes oficiais de inventário (AjusteInventarioOficial) no
# mesmo recorte - o resultado_valor que já existia em
# dashboard/fluxo-inventario-totais, também agora filtrável.
#
# O resumo narrado distingue, dentre as Divergências já RESOLVIDAS no
# recorte, quais foram encerradas como AJUSTE DE PROCESSO (erro de
# cadastro/contagem, timing de nota fiscal ou transferência, produção não
# encerrada etc. - catálogo em hipoteses_config.py) - ou seja, NÃO são
# perda real de estoque, só uma correção de dado/processo - das que foram
# CONFIRMADAS como perda real (avaria ou perda não identificada). Essa
# distinção é o que o Maurício pediu explicitamente: "pontuando casos
# justificados e solucionados como um não passivo real, apenas um ajuste
# de processo".
# ---------------------------------------------------------------------------

HIPOTESES_AJUSTE_PROCESSO = {
    "Transferencia_Pendente", "Consumo_Parcial_OP", "Pendencia_Faturamento", "Erro_Operacional",
    "Erro_Cadastro", "Falha_Inventario", "Producao_Nao_Encerrada", "Ajuste_Manual_Incorreto",
    "Movimentacao_Duplicada", "Conversao_Unidade_Incorreta", "Erro_Fiscal",
    "Divergencia_Ficha_Tecnica", "Pedido_Compra_Pendente", "Sem_Divergencia_Real",
}
HIPOTESES_PERDA_REAL = {"Avaria_Perda", "Perda_Nao_Identificada"}
# "Outros_Nao_Categorizado" fica de fora dos dois de propósito - não dá pra
# afirmar se é perda real ou ajuste de processo só pelo código da hipótese,
# então entra em "não_classificado" no resumo, não em nenhum dos dois lados.


def _parse_data_iso(valor: str | None, nome_campo: str):
    """Converte uma data YYYY-MM-DD vinda de query string, ou levanta 400 com
    mensagem clara se o formato vier errado - mesmo comportamento que já
    existia (repetido) em cada endpoint filtrável; centralizado aqui pra
    todos os endpoints do painel usarem igual."""
    if not valor:
        return None
    try:
        return date.fromisoformat(valor)
    except ValueError:
        raise HTTPException(400, f"{nome_campo} inválida (esperado YYYY-MM-DD): {valor}")


def _parse_motivos(motivo: str | None) -> list[str] | None:
    """Mesma convenção usada em todo o painel: motivo pode ser um valor único
    ou uma lista separada por vírgula (clique no segmento "Outros")."""
    return [m.strip() for m in motivo.split(",") if m.strip()] if motivo else None


def _data_no_periodo(d, ano: int | None, mes: int | None, data_inicio, data_fim) -> bool:
    """True se `d` (um date, pode ser None) cai dentro do recorte pedido.
    Sem NENHUM filtro de data informado, considera tudo (não teria como
    filtrar por data ausente)."""
    if ano is None and mes is None and data_inicio is None and data_fim is None:
        return True
    if not d:
        return False
    if data_inicio and d < data_inicio:
        return False
    if data_fim and d > data_fim:
        return False
    if ano is not None and d.year != ano:
        return False
    if mes is not None and d.month != mes:
        return False
    return True


@router.get("/dashboard/resumo-executivo/filtros")
def resumo_executivo_filtros(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Opções pros dropdowns de filtro da tela Mapeamento de Passivos -
    ano, almoxarifado e motivo, tirados dos valores que REALMENTE existem
    em baixas_operacionais (não do catálogo estático), pra nunca mostrar
    uma opção que não devolve nada."""
    datas = [d for (d,) in db.query(models.BaixaOperacional.data_baixa).distinct() if d]
    anos = sorted({d.year for d in datas}, reverse=True)
    almoxarifados = sorted({a for (a,) in db.query(models.BaixaOperacional.almoxarifado).distinct() if a})
    motivos = sorted({m for (m,) in db.query(models.BaixaOperacional.motivo_baixa_bruto).distinct() if m})
    return {"anos": anos, "meses": list(range(1, 13)), "almoxarifados": almoxarifados, "motivos": motivos}


def _montar_resumo_narrado(passivos, resultado_inventario, divergencias_resolvidas, ano, mes, di, df, almoxarifado, motivos_lista) -> str:
    filtro_txt = []
    if di or df:
        filtro_txt.append(f"entre {di.isoformat() if di else '...'} e {df.isoformat() if df else '...'}")
    elif ano and mes:
        filtro_txt.append(f"em {mes:02d}/{ano}")
    elif ano:
        filtro_txt.append(f"no ano {ano}")
    elif mes:
        filtro_txt.append(f"no mês {mes:02d} (todos os anos)")
    if almoxarifado:
        filtro_txt.append(f"almoxarifado {almoxarifado}")
    if motivos_lista:
        filtro_txt.append(f"motivo(s) {', '.join(motivos_lista)}")
    escopo = " · ".join(filtro_txt) if filtro_txt else "todo o histórico, sem filtro"

    partes = [f"Recorte considerado: {escopo}."]

    cat = passivos["por_categoria"]
    partes.append(
        f"Passivos: R$ {passivos['valor']:.2f} aprovados em {passivos['quantidade']} baixa(s) operacional(is). "
        f"Dessas, {cat['inventario_mensal']['quantidade']} já foram mapeadas via inventário mensal "
        f"(R$ {cat['inventario_mensal']['valor']:.2f}), {cat['movimentacao_diaria']['quantidade']} via "
        f"movimentação diária (R$ {cat['movimentacao_diaria']['valor']:.2f}), e "
        f"{cat['aguardando_divergencia']['quantidade']} continuam aprovadas mas aguardando cruzamento com uma "
        f"divergência (R$ {cat['aguardando_divergencia']['valor']:.2f})."
    )

    ri = resultado_inventario
    partes.append(
        f"Resultado de Inventário Acumulado: Entradas R$ {ri['entradas_valor']:.2f} − Saídas "
        f"R$ {ri['saidas_valor']:.2f} = R$ {ri['resultado_valor']:.2f} ({ri['entradas_qtd']} ajuste(s) de entrada, "
        f"{ri['saidas_qtd']} de saída, neste recorte)."
    )
    if ri["entradas_qtd"] + ri["saidas_qtd"] > 0:
        perda_fisica = -ri["resultado_valor"]
        diferenca = round(perda_fisica - passivos["valor"], 2)
        if abs(diferenca) < 1:
            partes.append("O valor aprovado como baixa bate de perto com a perda física medida pelo inventário neste recorte.")
        elif diferenca > 0:
            partes.append(
                f"Atenção: a perda física medida pelo inventário (R$ {perda_fisica:.2f}) é maior que o valor já "
                f"aprovado como baixa operacional (R$ {passivos['valor']:.2f}) neste recorte - uma diferença de "
                f"R$ {diferenca:.2f} de perda física ainda sem uma baixa operacional aprovada que a explique."
            )
        else:
            partes.append(
                f"O valor aprovado como baixa operacional (R$ {passivos['valor']:.2f}) é maior que a perda física "
                f"líquida medida pelo inventário (R$ {perda_fisica:.2f}) neste recorte - diferença de "
                f"R$ {abs(diferenca):.2f}, possivelmente baixas de um período que o inventário físico deste "
                f"recorte ainda não cobriu."
            )
    else:
        partes.append("Não há ajuste de inventário oficial registrado nesse recorte pra comparar com o valor de baixa aprovado.")

    dp, dr, dn = divergencias_resolvidas["ajuste_processo"], divergencias_resolvidas["perda_real"], divergencias_resolvidas["nao_classificado"]
    total_resolvidas = dp["quantidade"] + dr["quantidade"] + dn["quantidade"]
    if total_resolvidas:
        trecho = (
            f"Das divergências já resolvidas neste recorte ({total_resolvidas} no total): {dp['quantidade']} "
            f"(R$ {dp['valor']:.2f}) foram justificadas como AJUSTE DE PROCESSO - erro de cadastro, contagem, "
            f"timing de nota fiscal/transferência, produção não encerrada etc. - ou seja, NÃO representam perda "
            f"real de estoque; {dr['quantidade']} (R$ {dr['valor']:.2f}) foram confirmadas como PERDA REAL "
            f"(avaria ou perda não identificada)"
        )
        trecho += f"; e {dn['quantidade']} (R$ {dn['valor']:.2f}) ficaram com hipótese não classificada nesse critério." if dn["quantidade"] else "."
        partes.append(trecho)
    else:
        partes.append("Nenhuma divergência resolvida encontrada nesse recorte até agora.")

    return " ".join(partes)


@router.get("/dashboard/resumo-executivo")
def resumo_executivo(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Único endpoint por trás da tela Mapeamento de Passivos (versão
    simplificada, 12/08/2026): devolve os 2 indicadores centrais (Passivos,
    Resultado de Inventário Acumulado) e o texto do resumo executivo do
    pop-up de duplo-clique, todos já filtrados pelos mesmos critérios
    (Data/Mês/Ano, Almoxarifado, Motivo) - ver bloco de comentário acima."""
    try:
        di = date.fromisoformat(data_inicio) if data_inicio else None
    except ValueError:
        raise HTTPException(400, f"data_inicio inválida (esperado YYYY-MM-DD): {data_inicio}")
    try:
        df = date.fromisoformat(data_fim) if data_fim else None
    except ValueError:
        raise HTTPException(400, f"data_fim inválida (esperado YYYY-MM-DD): {data_fim}")
    motivos_lista = [m.strip() for m in motivo.split(",") if m.strip()] if motivo else None

    # --- Passivos (BaixaOperacional aprovada) ---
    q = db.query(models.BaixaOperacional)
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivos_lista:
        q = q.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    baixas_no_filtro = [b for b in q.all() if _data_no_periodo(b.data_baixa, ano, mes, di, df)]
    aprovadas = [b for b in baixas_no_filtro if b.status_fluxo == "APROVADA"]
    divergencias_por_id = _mapa_divergencias_das_baixas(db, aprovadas)
    por_categoria = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in aprovadas:
        c = _categoria_mapeamento(b, divergencias_por_id)
        por_categoria[c]["quantidade"] += 1
        por_categoria[c]["valor"] += b.valor_total or 0
    passivos = {
        "valor": round(sum(b.valor_total or 0 for b in aprovadas), 2),
        "quantidade": len(aprovadas),
        "total_no_filtro": len(baixas_no_filtro),
        "por_categoria": {
            chave: {"label": rotulo, "quantidade": por_categoria[chave]["quantidade"], "valor": round(por_categoria[chave]["valor"], 2)}
            for chave, rotulo in CATEGORIA_MAPEAMENTO_LABELS.items()
        },
    }

    # --- Resultado de Inventário Acumulado (AjusteInventarioOficial) ---
    qa = db.query(models.AjusteInventarioOficial).filter(models.AjusteInventarioOficial.conta_como_ajuste_inventario.is_(True))
    if almoxarifado:
        qa = qa.filter(models.AjusteInventarioOficial.almoxarifado == almoxarifado)
    ajustes = [a for a in qa.all() if a.ajuste_qtd and _data_no_periodo(a.dt_invent, ano, mes, di, df)]
    entradas_valor = round(sum(abs(a.valor_total or 0) for a in ajustes if a.ajuste_qtd > 0), 2)
    saidas_valor = round(sum(abs(a.valor_total or 0) for a in ajustes if a.ajuste_qtd < 0), 2)
    resultado_inventario = {
        "entradas_valor": entradas_valor, "saidas_valor": saidas_valor,
        "resultado_valor": round(entradas_valor - saidas_valor, 2),
        "entradas_qtd": sum(1 for a in ajustes if a.ajuste_qtd > 0),
        "saidas_qtd": sum(1 for a in ajustes if a.ajuste_qtd < 0),
    }

    # --- Divergências resolvidas: ajuste de processo vs perda real ---
    qd = db.query(models.Divergencia).filter(models.Divergencia.status == "Resolvida")
    if almoxarifado:
        qd = qd.filter(models.Divergencia.almoxarifado == almoxarifado)
    divergencias = [d for d in qd.all() if _data_no_periodo(d.data_deteccao, ano, mes, di, df)]
    ajuste_processo = [d for d in divergencias if d.hipotese_confirmada in HIPOTESES_AJUSTE_PROCESSO]
    perda_real = [d for d in divergencias if d.hipotese_confirmada in HIPOTESES_PERDA_REAL]
    ids_classificados = {d.id for d in ajuste_processo} | {d.id for d in perda_real}
    nao_classificado = [d for d in divergencias if d.id not in ids_classificados]
    divergencias_resolvidas = {
        "ajuste_processo": {"quantidade": len(ajuste_processo), "valor": round(sum(d.valor_estimado or 0 for d in ajuste_processo), 2)},
        "perda_real": {"quantidade": len(perda_real), "valor": round(sum(d.valor_estimado or 0 for d in perda_real), 2)},
        "nao_classificado": {"quantidade": len(nao_classificado), "valor": round(sum(d.valor_estimado or 0 for d in nao_classificado), 2)},
    }

    resumo_narrado = _montar_resumo_narrado(
        passivos, resultado_inventario, divergencias_resolvidas, ano, mes, di, df, almoxarifado, motivos_lista
    )

    return {
        "filtros_aplicados": {
            "ano": ano, "mes": mes, "data_inicio": data_inicio, "data_fim": data_fim,
            "almoxarifado": almoxarifado, "motivos": motivos_lista,
        },
        "passivos": passivos,
        "resultado_inventario": resultado_inventario,
        "divergencias_resolvidas": divergencias_resolvidas,
        "resumo_narrado": resumo_narrado,
    }


# ---------------------------------------------------------------------------
# Segundo indicador + exportação em Excel (13/08/2026) - pedido do Maurício:
# (1) uma quebra por ALMOXARIFADO (em vez de por mês) do mesmo par
#     Passivos x Resultado de Inventário do Resumo Executivo, pra achar onde
#     estão concentradas as maiores perdas;
# (2) exportar o recorte atual pra Excel, com Inventário/Passivos/Acumulado/
#     Resumo em abas separadas.
# Os dois reaproveitam os mesmos helpers de filtro (_data_no_periodo,
# _parse_data_iso, _parse_motivos) do Resumo Executivo, pro recorte ser
# idêntico em toda a tela.
# ---------------------------------------------------------------------------

def _dados_inventario_filtrados(db: Session, ano, mes, data_inicio, data_fim, almoxarifado) -> list:
    q = db.query(models.AjusteInventarioOficial).filter(models.AjusteInventarioOficial.conta_como_ajuste_inventario.is_(True))
    if almoxarifado:
        q = q.filter(models.AjusteInventarioOficial.almoxarifado == almoxarifado)
    return [a for a in q.all() if a.ajuste_qtd and _data_no_periodo(a.dt_invent, ano, mes, data_inicio, data_fim)]


def _dados_passivos_filtrados(db: Session, ano, mes, data_inicio, data_fim, almoxarifado, motivos_lista) -> list:
    q = db.query(models.BaixaOperacional)
    if almoxarifado:
        q = q.filter(models.BaixaOperacional.almoxarifado == almoxarifado)
    if motivos_lista:
        q = q.filter(models.BaixaOperacional.motivo_baixa_bruto.in_(motivos_lista))
    return [b for b in q.all() if _data_no_periodo(b.data_baixa, ano, mes, data_inicio, data_fim)]


@router.get("/dashboard/resultado-por-almoxarifado")
def dashboard_resultado_por_almoxarifado(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Segundo indicador do Mapeamento de Passivos: a mesma quebra Passivos
    x Resultado de Inventário do Resumo Executivo, só que por ALMOXARIFADO
    em vez de por mês - "onde estão as maiores perdas". Uma linha por
    almoxarifado, ordenada do maior pro menor "valor acumulado do período"
    (Passivos + Resultado de Inventário em valor absoluto) - o gráfico de
    barras horizontal empilhado que o Maurício pediu."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    motivos_lista = _parse_motivos(motivo)

    aprovadas = [
        b for b in _dados_passivos_filtrados(db, ano, mes, di, df, almoxarifado, motivos_lista)
        if b.status_fluxo == "APROVADA"
    ]
    passivos_por_almox = defaultdict(float)
    for b in aprovadas:
        passivos_por_almox[b.almoxarifado or "Não informado"] += b.valor_total or 0

    ajustes = _dados_inventario_filtrados(db, ano, mes, di, df, almoxarifado)
    resultado_por_almox = defaultdict(float)
    for a in ajustes:
        # valor_total já vem com o mesmo sinal do ajuste_qtd (positivo=entrada, negativo=saída) -
        # somar direto já dá o resultado NETO (entradas - saídas) por almoxarifado.
        resultado_por_almox[a.almoxarifado or "Não informado"] += a.valor_total or 0

    almoxarifados = sorted(set(passivos_por_almox) | set(resultado_por_almox))
    linhas = []
    for almox in almoxarifados:
        passivos_valor = round(passivos_por_almox.get(almox, 0.0), 2)
        resultado_inventario_valor = round(resultado_por_almox.get(almox, 0.0), 2)
        inventario_valor_abs = round(abs(resultado_inventario_valor), 2)
        linhas.append({
            "almoxarifado": almox,
            "passivos_valor": passivos_valor,
            "resultado_inventario_valor": resultado_inventario_valor,
            "inventario_valor_abs": inventario_valor_abs,
            "valor_acumulado": round(passivos_valor + inventario_valor_abs, 2),
        })
    linhas.sort(key=lambda x: x["valor_acumulado"], reverse=True)
    return linhas


@router.get("/dashboard/top-10-movimentos")
def dashboard_top_10_movimentos(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """As 10 maiores movimentações do recorte atual (13/08/2026), misturando
    Passivos aprovados e Ajustes de Inventário numa lista só, ordenada por
    valor absoluto - complementa o "Resultado por Almoxarifado" (que soma
    por almoxarifado): aqui é o detalhe linha a linha das maiores baixas
    específicas do período, de onde vier. `tipo` distingue "passivo" de
    "inventario" - o front usa isso pra decidir que ação de justificativa
    oferecer em cada linha (ver abrirModalJustificativaPorItem)."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df = _parse_data_iso(data_fim, "data_fim")
    motivos_lista = _parse_motivos(motivo)

    aprovadas = [
        b for b in _dados_passivos_filtrados(db, ano, mes, di, df, almoxarifado, motivos_lista)
        if b.status_fluxo == "APROVADA"
    ]
    ajustes = _dados_inventario_filtrados(db, ano, mes, di, df, almoxarifado)

    skus = {b.sku for b in aprovadas if b.sku} | {a.sku for a in ajustes if a.sku}
    descricoes = {p.sku: p.descricao for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus)).all()}

    itens = []
    for b in aprovadas:
        itens.append({
            "tipo": "passivo", "id": b.id, "sku": b.sku, "descricao_produto": descricoes.get(b.sku),
            "almoxarifado": b.almoxarifado, "data": str(b.data_baixa) if b.data_baixa else None,
            # Passivo sempre entra como valor NEGATIVO nesta tabela - diferente do Ajuste de
            # Inventário (que pode ser entrada ou saída), um Passivo aprovado é sempre uma perda
            # que reduz o resultado do período, então o sinal aqui não depende de como o
            # valor_total foi gravado (pode vir positivo do Lovable).
            "quantidade": b.quantidade, "valor": round(abs(b.valor_total or 0), 2), "valor_com_sinal": round(-abs(b.valor_total or 0), 2),
            "motivo": b.motivo_baixa_bruto, "status_fluxo": b.status_fluxo,
            "divergencia_vinculada_id": b.divergencia_vinculada_id,
            "id_lote": None, "qtd_sistema": None, "qtd_contagem": None, "direcao": None,
        })
    for a in ajustes:
        itens.append({
            "tipo": "inventario", "id": a.id, "sku": a.sku, "descricao_produto": a.descricao_produto or descricoes.get(a.sku),
            "almoxarifado": a.almoxarifado, "data": str(a.dt_invent) if a.dt_invent else None,
            "quantidade": a.ajuste_qtd, "valor": round(abs(a.valor_total or 0), 2), "valor_com_sinal": round(a.valor_total or 0, 2),
            "motivo": None, "status_fluxo": None, "divergencia_vinculada_id": None,
            "id_lote": a.id_lote, "qtd_sistema": a.qtd_sistema, "qtd_contagem": a.qtd_contagem,
            "direcao": _direcao_ajuste(a),
        })
    itens.sort(key=lambda x: x["valor"], reverse=True)
    top10 = itens[:10]

    ids_passivo = {i["id"] for i in top10 if i["tipo"] == "passivo"}
    ids_inventario = {i["id"] for i in top10 if i["tipo"] == "inventario"}
    baixas_com_justificativa = set()
    ajustes_com_justificativa = set()
    if ids_passivo:
        baixas_com_justificativa = {
            j.baixa_operacional_id
            for j in db.query(models.JustificativaAjusteInventario)
            .filter(models.JustificativaAjusteInventario.baixa_operacional_id.in_(ids_passivo))
            .all()
        }
    if ids_inventario:
        ajustes_com_justificativa = {
            j.ajuste_id
            for j in db.query(models.JustificativaAjusteInventario)
            .filter(models.JustificativaAjusteInventario.ajuste_id.in_(ids_inventario))
            .all()
        }

    for item in top10:
        if item["tipo"] == "passivo":
            item["tem_justificativa"] = item["id"] in baixas_com_justificativa
        else:
            item["tem_justificativa"] = item["id"] in ajustes_com_justificativa

    return top10


@router.get("/dashboard/exportar-excel")
def dashboard_passivos_exportar_excel(
    ano: int | None = None,
    mes: int | None = Query(None, ge=1, le=12),
    data_inicio: str | None = Query(None, description="YYYY-MM-DD"),
    data_fim: str | None = Query(None, description="YYYY-MM-DD"),
    almoxarifado: str | None = None,
    motivo: str | None = Query(None, description="motivo_baixa_bruto exato, ou lista separada por vírgula"),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Exporta o recorte atual do Mapeamento de Passivos (mesmos filtros da
    tela) pra um Excel com 4 abas: "Inventário" (AjusteInventarioOficial
    contado como ajuste de inventário), "Passivos" (BaixaOperacional, de
    qualquer status - a coluna Status deixa filtrar na própria planilha),
    "Acumulado" (as duas juntas, colunas normalizadas, pra somar tudo numa
    tabela só) e "Resumo (Dinâmica)" com os totais do recorte, por
    almoxarifado e por motivo - a "dinâmica com os principais resumos"
    pedida pelo Maurício."""
    di = _parse_data_iso(data_inicio, "data_inicio")
    df_ = _parse_data_iso(data_fim, "data_fim")
    motivos_lista = _parse_motivos(motivo)

    ajustes = _dados_inventario_filtrados(db, ano, mes, di, df_, almoxarifado)
    baixas = _dados_passivos_filtrados(db, ano, mes, di, df_, almoxarifado, motivos_lista)

    skus_relevantes = {a.sku for a in ajustes if a.sku} | {b.sku for b in baixas if b.sku}
    descricoes = {p.sku: p.descricao for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus_relevantes)).all()}

    df_inventario = pd.DataFrame([{
        "SKU": a.sku, "Descrição": a.descricao_produto or descricoes.get(a.sku),
        "Almoxarifado": a.almoxarifado, "Categoria": a.categoria_produto,
        "ID Lote": a.id_lote, "ID Inventário": a.id_invent, "Data do Inventário": a.dt_invent,
        "Qtd Sistema": a.qtd_sistema, "Qtd Contagem": a.qtd_contagem, "Ajuste Qtd": a.ajuste_qtd,
        "Custo Unitário": a.custo_unitario, "Valor Total": a.valor_total, "Observação": a.observacao,
    } for a in ajustes])

    df_passivos = pd.DataFrame([{
        "ID": b.id, "SKU": b.sku, "Descrição": descricoes.get(b.sku),
        "Almoxarifado": b.almoxarifado, "Motivo": b.motivo_baixa_bruto, "Hipótese Aplicada": b.hipotese_aplicada,
        "Quantidade": b.quantidade, "Valor Total": b.valor_total, "Status": b.status_fluxo,
        "Solicitante": b.solicitante_nome, "Data da Baixa": b.data_baixa,
        "Divergência Vinculada": b.divergencia_vinculada_id,
    } for b in baixas])

    linhas_acumulado = [
        {
            "Tipo": "Inventário", "SKU": a.sku, "Descrição": a.descricao_produto or descricoes.get(a.sku),
            "Almoxarifado": a.almoxarifado, "Data": a.dt_invent, "Quantidade": a.ajuste_qtd,
            "Valor": a.valor_total, "Detalhe": a.observacao or "",
        }
        for a in ajustes
    ] + [
        {
            "Tipo": "Passivo", "SKU": b.sku, "Descrição": descricoes.get(b.sku),
            "Almoxarifado": b.almoxarifado, "Data": b.data_baixa, "Quantidade": b.quantidade,
            "Valor": b.valor_total, "Detalhe": f"{b.motivo_baixa_bruto or ''} · {b.status_fluxo or ''}".strip(" ·"),
        }
        for b in baixas
    ]
    df_acumulado = pd.DataFrame(linhas_acumulado)
    if not df_acumulado.empty:
        df_acumulado = df_acumulado.sort_values("Data", ascending=False, na_position="last")

    # --- Resumo (dinâmica): totais gerais + por almoxarifado + por motivo ---
    aprovadas = [b for b in baixas if b.status_fluxo == "APROVADA"]
    passivos_por_almox = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in aprovadas:
        chave = b.almoxarifado or "Não informado"
        passivos_por_almox[chave]["quantidade"] += 1
        passivos_por_almox[chave]["valor"] += b.valor_total or 0
    resultado_por_almox = defaultdict(float)
    for a in ajustes:
        resultado_por_almox[a.almoxarifado or "Não informado"] += a.valor_total or 0

    linhas_resumo_almox = []
    for almox in sorted(set(passivos_por_almox) | set(resultado_por_almox)):
        p = passivos_por_almox.get(almox, {"quantidade": 0, "valor": 0.0})
        resultado_valor = round(resultado_por_almox.get(almox, 0.0), 2)
        linhas_resumo_almox.append({
            "Almoxarifado": almox, "Passivos (Qtd)": p["quantidade"], "Passivos (R$)": round(p["valor"], 2),
            "Resultado Inventário (R$)": resultado_valor,
            "Valor Acumulado (R$)": round(p["valor"] + abs(resultado_valor), 2),
        })
    linhas_resumo_almox.sort(key=lambda x: x["Valor Acumulado (R$)"], reverse=True)
    df_resumo_almox = pd.DataFrame(linhas_resumo_almox)
    if not df_resumo_almox.empty:
        df_resumo_almox = pd.concat([df_resumo_almox, pd.DataFrame([{
            "Almoxarifado": "TOTAL",
            "Passivos (Qtd)": int(df_resumo_almox["Passivos (Qtd)"].sum()),
            "Passivos (R$)": round(df_resumo_almox["Passivos (R$)"].sum(), 2),
            "Resultado Inventário (R$)": round(df_resumo_almox["Resultado Inventário (R$)"].sum(), 2),
            "Valor Acumulado (R$)": round(df_resumo_almox["Valor Acumulado (R$)"].sum(), 2),
        }])], ignore_index=True)

    por_motivo = defaultdict(lambda: {"quantidade": 0, "valor": 0.0})
    for b in aprovadas:
        chave = b.motivo_baixa_bruto or "Não informado"
        por_motivo[chave]["quantidade"] += 1
        por_motivo[chave]["valor"] += b.valor_total or 0
    linhas_resumo_motivo = sorted(
        [{"Motivo": motivo_k, "Quantidade": v["quantidade"], "Valor (R$)": round(v["valor"], 2)} for motivo_k, v in por_motivo.items()],
        key=lambda x: abs(x["Valor (R$)"]), reverse=True,
    )
    df_resumo_motivo = pd.DataFrame(linhas_resumo_motivo)

    total_passivos_valor = round(sum(b.valor_total or 0 for b in aprovadas), 2)
    total_inventario_valor = round(sum(a.valor_total or 0 for a in ajustes), 2)
    df_resumo_totais = pd.DataFrame([
        {"Indicador": "Passivos aprovados no recorte (R$)", "Valor": total_passivos_valor},
        {"Indicador": "Resultado de Inventário Acumulado no recorte (R$)", "Valor": total_inventario_valor},
        {"Indicador": "Valor acumulado do recorte - Passivos + |Resultado Inventário| (R$)", "Valor": round(total_passivos_valor + abs(total_inventario_valor), 2)},
        {"Indicador": "Total de baixas aprovadas no recorte", "Valor": len(aprovadas)},
        {"Indicador": "Total de ajustes de inventário no recorte", "Valor": len(ajustes)},
    ])

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        (df_inventario if not df_inventario.empty else pd.DataFrame([{"Aviso": "Nenhum ajuste de inventário encontrado neste recorte."}])).to_excel(
            writer, sheet_name="Inventário", index=False)
        (df_passivos if not df_passivos.empty else pd.DataFrame([{"Aviso": "Nenhuma baixa encontrada neste recorte."}])).to_excel(
            writer, sheet_name="Passivos", index=False)
        (df_acumulado if not df_acumulado.empty else pd.DataFrame([{"Aviso": "Nenhum dado encontrado neste recorte."}])).to_excel(
            writer, sheet_name="Acumulado", index=False)

        linha_atual = 0
        df_resumo_totais.to_excel(writer, sheet_name="Resumo (Dinâmica)", index=False, startrow=linha_atual)
        linha_atual += len(df_resumo_totais) + 3
        (df_resumo_almox if not df_resumo_almox.empty else pd.DataFrame([{"Almoxarifado": "Sem dados neste recorte."}])).to_excel(
            writer, sheet_name="Resumo (Dinâmica)", index=False, startrow=linha_atual)
        linha_atual += max(len(df_resumo_almox), 1) + 3
        (df_resumo_motivo if not df_resumo_motivo.empty else pd.DataFrame([{"Motivo": "Sem dados neste recorte."}])).to_excel(
            writer, sheet_name="Resumo (Dinâmica)", index=False, startrow=linha_atual)

        # cabeçalho em negrito (linha 1 das 3 primeiras abas + início de cada bloco na aba Resumo)
        for nome_aba in ("Inventário", "Passivos", "Acumulado"):
            for celula in writer.sheets[nome_aba][1]:
                celula.font = Font(bold=True)
        ws_resumo = writer.sheets["Resumo (Dinâmica)"]
        for linha_planilha in ws_resumo.iter_rows():
            if linha_planilha[0].value in ("Indicador", "Almoxarifado", "Motivo"):
                for celula in linha_planilha:
                    celula.font = Font(bold=True)

        # largura de coluna proporcional ao conteúdo, em toda aba
        for nome_aba in writer.sheets:
            ws = writer.sheets[nome_aba]
            for coluna in ws.columns:
                maior = max((len(str(c.value)) for c in coluna if c.value is not None), default=10)
                ws.column_dimensions[coluna[0].column_letter].width = min(max(maior + 2, 10), 40)

    buffer.seek(0)
    nome_arquivo = "mapeamento_passivos" + (f"_{almoxarifado}" if almoxarifado else "") + ".xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )
