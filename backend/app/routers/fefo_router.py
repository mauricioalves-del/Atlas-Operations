"""
Módulo FEFO (First-Expired-First-Out) - detecção de quebras na
movimentação de saída da Fábrica (18/08/2026). Motor de cálculo nativo
(20/08/2026) baseado em lote movimentado (ver app/fefo.py e
claude/checagens-fefo-heuristica-quebrada.md pra histórico/regra completa);
models.ChecagemFefo/Transferencia (heurística antiga) e a "Auditoria FEFO
importada" (relatórios já prontos do André) continuam existindo como
features separadas - ver docstrings correspondentes.
"""
import io
from datetime import date
from typing import Optional
import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import obter_usuario_atual, requer_papel
from ..audit import registrar_log
from .. import fefo

router = APIRouter(prefix="/fefo", tags=["fefo"])


# ═════════════════════════════════════════════════════════════════════════
# Motor NATIVO de checagem de FEFO por lote movimentado (20/08/2026) - ver
# docstring da seção equivalente em app/fefo.py e
# claude/checagens-fefo-heuristica-quebrada.md no Atlas Operations pro
# porquê disso substituir o cálculo antigo baseado em Transferencia/
# ChecagemFefo (desativado, 89,85% de falso positivo).
# ═════════════════════════════════════════════════════════════════════════

@router.post("/movimentacao/importar")
async def importar_movimentacao_lote(
    arquivos: list[UploadFile] = File(...),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa um ou mais Excels de movimentação bruta POR LOTE
    ("Movimentação - Lt.xlsx", a mesma exportação que o André já usa no
    processo dele - ver Auditar_FEFO.ipynb, ler_movimentacao()). Um arquivo
    pode trazer vários dias de uma vez; reimportar substitui só os dias
    presentes no arquivo novo. Recalcula a checagem de FEFO automaticamente
    ao final de cada importação (além do recálculo diário automático em
    background - ver scheduler.py)."""
    resultados = []
    for arquivo in arquivos:
        conteudo = await arquivo.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
        except Exception as e:
            resultados.append({"arquivo": arquivo.filename, "erro": f"Não consegui abrir o Excel: {e}"})
            continue
        ws = wb[wb.sheetnames[0]]
        linhas_brutas = list(ws.iter_rows(values_only=True))
        if not linhas_brutas:
            resultados.append({"arquivo": arquivo.filename, "erro": "Planilha vazia."})
            continue
        cabecalho = [str(c).strip() if c is not None else "" for c in linhas_brutas[0]]
        faltando = [c for c in fefo.COLUNAS_MOVIMENTACAO_LOTE_DIARIA if c not in cabecalho]
        if faltando:
            resultados.append({
                "arquivo": arquivo.filename,
                "erro": f"Colunas obrigatórias não encontradas: {faltando}. Cabeçalho encontrado: {cabecalho}",
            })
            continue
        linhas = [dict(zip(cabecalho, linha)) for linha in linhas_brutas[1:] if any(v is not None for v in linha)]
        resultado = fefo.importar_movimentacao_lote_diaria(db, linhas, arquivo.filename, usuario.username)
        resultados.append({"arquivo": arquivo.filename, **resultado})

    db.commit()
    registrar_log(db, usuario.username, "importar_movimentacao_lote", detalhes={"arquivos": [r["arquivo"] for r in resultados]})
    total_importadas = sum(r.get("linhas_importadas", 0) for r in resultados)
    total_quebras = sum(r.get("quebras_detectadas", 0) for r in resultados)
    total_erros = sum(1 for r in resultados if "erro" in r)
    return {"arquivos_processados": len(resultados), "arquivos_com_erro": total_erros,
            "linhas_importadas": total_importadas, "quebras_detectadas": total_quebras, "detalhe": resultados}


@router.post("/recalcular")
def recalcular(usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Roda o motor nativo de checagem de FEFO (compara o lote que de fato
    saiu da Fábrica contra os lotes do mesmo SKU que ficaram lá) pra todo o
    histórico de movimentação por lote importado, e atualiza o resultado
    guardado - chamar depois de reimportar a planilha de lotes
    (Lote_Sistema), mesmo sem reimportar movimentação nova."""
    resultado = fefo.recalcular_quebra_fefo_nativa(db)
    registrar_log(db, usuario.username, "recalcular_fefo", detalhes=resultado)
    db.commit()
    return resultado


@router.get("/checagens")
def listar_checagens(
    apenas_quebras: bool = False,
    sku: Optional[str] = None,
    almoxarifado_destino: Optional[str] = None,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Detalhamento do motor nativo (ChecagemFefoMovimento) - um registro
    por movimento de saída da Fábrica avaliado, com o lote que de fato
    saiu e, quando houve quebra, o lote mais antigo que deveria ter saído."""
    q = db.query(models.ChecagemFefoMovimento)
    if apenas_quebras:
        q = q.filter(models.ChecagemFefoMovimento.quebra_fefo.is_(True))
    if sku:
        q = q.filter(models.ChecagemFefoMovimento.sku == sku)
    if almoxarifado_destino:
        q = q.filter(models.ChecagemFefoMovimento.almoxarifado_destino == almoxarifado_destino)
    registros = q.order_by(models.ChecagemFefoMovimento.data.desc(), models.ChecagemFefoMovimento.id.desc()).limit(2000).all()
    return [
        {
            "id": r.id, "data": str(r.data), "sku": r.sku, "descricao_produto": r.descricao_produto,
            "movimento": r.movimento, "almoxarifado_destino": r.almoxarifado_destino,
            "lote_movimentado": r.lote_movimentado, "qtd_lote_movimentado": r.qtd_lote_movimentado,
            "validade_lote_movimentado": str(r.validade_lote_movimentado) if r.validade_lote_movimentado else None,
            "quebra_fefo": r.quebra_fefo, "status": r.status,
            "lote_mais_antigo_disponivel": r.lote_mais_antigo_disponivel,
            "qtd_lote_mais_antigo_disponivel": r.qtd_lote_mais_antigo_disponivel,
            "validade_mais_antiga_disponivel": str(r.validade_mais_antiga_disponivel) if r.validade_mais_antiga_disponivel else None,
        }
        for r in registros
    ]


@router.get("/dashboard/resumo")
def dashboard_resumo(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Total de movimentos avaliados (saídas da Fábrica com checagem já
    calculada pelo motor nativo), quantos quebraram o critério de FEFO,
    taxa de quebra, e os SKUs/destinos mais frequentes nas quebras - pro
    indicador do MBR. Rodar POST /fefo/recalcular antes se os dados de
    movimentação ou de lotes tiverem mudado desde a última checagem."""
    return fefo.calcular_resumo_checagem_fefo_movimento(db)


# ═════════════════════════════════════════════════════════════════════════
# Auditoria FEFO importada (20/08/2026) - ver docstring de
# models.AuditoriaFefo e da seção equivalente em app/fefo.py.
# ═════════════════════════════════════════════════════════════════════════

@router.post("/auditoria/importar")
async def importar_auditoria_fefo(
    arquivos: list[UploadFile] = File(...),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa um ou mais Excels de auditoria FEFO diária (o arquivo que o
    processo do estagiário já gera todo dia - ver Auditar_FEFO.ipynb dele),
    lendo a aba 'Todas as Movimentações'. Pode subir vários arquivos de uma
    vez (ex: pra consolidar o histórico acumulado). Cada arquivo é um dia -
    reimportar o mesmo dia substitui as linhas daquele dia, não duplica."""
    resultados = []
    for arquivo in arquivos:
        conteudo = await arquivo.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
        except Exception as e:
            resultados.append({"arquivo": arquivo.filename, "erro": f"Não consegui abrir o Excel: {e}"})
            continue
        if fefo.ABA_AUDITORIA_FEFO_DIARIA not in wb.sheetnames:
            resultados.append({
                "arquivo": arquivo.filename,
                "erro": f"Aba '{fefo.ABA_AUDITORIA_FEFO_DIARIA}' não encontrada. Abas disponíveis: {wb.sheetnames}",
            })
            continue
        ws = wb[fefo.ABA_AUDITORIA_FEFO_DIARIA]
        linhas_brutas = list(ws.iter_rows(values_only=True))
        if not linhas_brutas:
            resultados.append({"arquivo": arquivo.filename, "erro": "Aba vazia."})
            continue
        cabecalho = [str(c).strip() if c is not None else "" for c in linhas_brutas[0]]
        faltando = [c for c in fefo.COLUNAS_AUDITORIA_FEFO_DIARIA if c not in cabecalho]
        if faltando:
            resultados.append({
                "arquivo": arquivo.filename,
                "erro": f"Colunas obrigatórias não encontradas: {faltando}. Cabeçalho encontrado: {cabecalho}",
            })
            continue
        linhas = [dict(zip(cabecalho, linha)) for linha in linhas_brutas[1:] if any(v is not None for v in linha)]
        resultado = fefo.importar_auditoria_fefo_diaria(db, linhas, arquivo.filename, usuario.username)
        resultados.append({"arquivo": arquivo.filename, **resultado})

    db.commit()
    registrar_log(db, usuario.username, "importar_auditoria_fefo", detalhes={"arquivos": [r["arquivo"] for r in resultados]})
    total_importadas = sum(r.get("linhas_importadas", 0) for r in resultados)
    total_quebras = sum(r.get("quebras_no_arquivo", 0) for r in resultados)
    total_erros = sum(1 for r in resultados if "erro" in r)
    return {"arquivos_processados": len(resultados), "arquivos_com_erro": total_erros,
            "linhas_importadas": total_importadas, "quebras_importadas": total_quebras, "detalhe": resultados}


@router.post("/auditoria/importar-consolidado")
async def importar_auditoria_fefo_consolidado(
    arquivo: UploadFile = File(...),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa o dashboard HTML consolidado que o estagiário já mantinha
    (Controle - FEFO.html, gerado pelo DashBoard_FEFO.ipynb dele) - só pra
    estender o histórico a dias sem o Excel de auditoria diária bruto. Ver
    fefo.importar_auditoria_fefo_consolidada pra regra de "dia já cobrido
    pela auditoria diária não é sobrescrito"."""
    conteudo = await arquivo.read()
    try:
        registros = fefo.extrair_registros_dashboard_consolidado(conteudo.decode("utf-8"))
    except Exception as e:
        raise HTTPException(400, f"Não consegui ler o dashboard consolidado: {e}")

    resultado = fefo.importar_auditoria_fefo_consolidada(db, registros, arquivo.filename, usuario.username)
    registrar_log(db, usuario.username, "importar_auditoria_fefo_consolidado",
                  detalhes={"arquivo": arquivo.filename, **resultado})
    db.commit()
    return {"arquivo": arquivo.filename, **resultado}


@router.get("/auditoria/resumo")
def resumo_auditoria_fefo(
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Agregados do histórico de auditoria FEFO importado (ver
    fefo.calcular_resumo_auditoria_fefo) - alimenta o painel nativo na tela
    FEFO, com o histórico real consolidado do processo do estagiário."""
    return fefo.calcular_resumo_auditoria_fefo(db, data_inicio, data_fim)


@router.get("/auditoria/registros")
def listar_auditoria_fefo(
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    apenas_quebras: bool = False,
    produto: Optional[str] = Query(None, description="Busca por SKU ou descrição (contém, case-insensitive)"),
    limite: int = Query(200, le=2000),
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Lista os movimentos importados, mais recentes primeiro - pra tabela
    de detalhamento da tela FEFO. Exclui 'Destino (não auditado)' por
    padrão (não é útil olhar individualmente), a menos que se queira tudo
    via outro filtro no futuro."""
    q = db.query(models.AuditoriaFefo).filter(models.AuditoriaFefo.status != fefo.STATUS_DESTINO_NAO_AUDITADO)
    if data_inicio:
        q = q.filter(models.AuditoriaFefo.data >= data_inicio)
    if data_fim:
        q = q.filter(models.AuditoriaFefo.data <= data_fim)
    if apenas_quebras:
        q = q.filter(models.AuditoriaFefo.quebra_fefo.is_(True))
    if produto:
        termo = f"%{produto}%"
        q = q.filter((models.AuditoriaFefo.sku.ilike(termo)) | (models.AuditoriaFefo.descricao_produto.ilike(termo)))
    registros = q.order_by(models.AuditoriaFefo.data.desc(), models.AuditoriaFefo.id.desc()).limit(limite).all()
    return [
        {
            "id": r.id, "data": str(r.data), "sku": r.sku, "descricao_produto": r.descricao_produto,
            "movimento": r.movimento, "almoxarifado": r.almoxarifado, "lote_movimentado": r.lote_movimentado,
            "qtd_lote_movimentado": r.qtd_lote_movimentado,
            "validade_lote_movimentado": str(r.validade_lote_movimentado) if r.validade_lote_movimentado else None,
            "quebra_fefo": r.quebra_fefo, "status": r.status,
            "lote_mais_antigo_disponivel": r.lote_mais_antigo_disponivel,
            "qtd_lote_mais_antigo_disponivel": r.qtd_lote_mais_antigo_disponivel,
            "validade_mais_antiga_disponivel": str(r.validade_mais_antiga_disponivel) if r.validade_mais_antiga_disponivel else None,
            "fonte": r.fonte,
        }
        for r in registros
    ]
