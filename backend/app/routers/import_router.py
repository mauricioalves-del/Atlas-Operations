import io
import csv as csv_module
from datetime import date, datetime
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.orm import Session
import openpyxl

from .. import models
from ..database import get_db
from ..csv_utils import parse_sku, parse_data, parse_decimal
from ..hipoteses_config import normalizar_almoxarifado
from ..investigation import investigar, reconciliar
from ..ml import predict as ml_predict
from ..deps import requer_papel, obter_usuario_atual
from ..audit import registrar_log

router = APIRouter(prefix="/importar", tags=["importacao"])

# Cabeçalhos esperados na aba "Movimentação" da planilha operacional real
# (arquivo tipo "Movimentados - Processo DD.MM.xlsm"). Se a planilha da sua
# empresa usar nomes um pouco diferentes, ajuste este de-para em vez de
# mudar a lógica de importação.
COLUNAS_EXCEL = {
    "grupo": "Grupo",
    "data": "Data",
    "sku": "Id_Produto",
    "unidade": "Unid.",
    "saldo_sistema": "Sistema",
    "entrada": "Entrada",
    "saida": "Saídas",
    "saldo_fisico": "Contagem",
    "obs": "Obs",
    "prejuizo": "Prejuízo",
}


def _categoria_do_sku(db: Session, sku: str, fallback: str | None = None) -> str:
    p = db.query(models.Produto).filter_by(sku=sku).first()
    if p and p.categoria_produto:
        return p.categoria_produto
    return fallback or "Desconhecido"


def _custo_unitario_do_sku(db: Session, sku: str) -> float | None:
    p = db.query(models.Produto).filter_by(sku=sku).first()
    return p.custo_unitario if p else None


def processar_linha_movimentacao(db: Session, row: dict, almoxarifado_forcado: str | None = None,
                                  categoria_fallback: str | None = None,
                                  observacao: str | None = None, prejuizo: bool = False,
                                  lote_id: int | None = None) -> str:
    """Recebe uma linha (sku, almoxarifado, data_movimento, entrada, saida,
    saldo_sistema, saldo_fisico, unidade) e decide: se saldo_sistema ==
    saldo_fisico, vai para o histórico como resolvido; senão, cria uma
    Divergencia nova e dispara investigação + ML + reconciliação.
    Retorna "historico" ou "divergencia".

    almoxarifado_forcado é usado pela importação de Excel, onde o arquivo
    não traz coluna de almoxarifado por linha (o arquivo já é de um
    almoxarifado só, escolhido no formulário de importação). lote_id
    marca de qual importação essa linha veio, pra dar pra desfazer depois."""
    sku = parse_sku(row["sku"])
    almoxarifado = (almoxarifado_forcado or row["almoxarifado"]).strip()
    data_mov = row["data_movimento"] if isinstance(row["data_movimento"], date) else parse_data(row["data_movimento"])
    entrada = parse_decimal(row.get("entrada", 0))
    saida = parse_decimal(row.get("saida", 0))
    saldo_sistema = parse_decimal(row["saldo_sistema"])
    saldo_fisico = parse_decimal(row["saldo_fisico"])
    unidade = row.get("unidade")
    divergencia_qtd = round(saldo_fisico - saldo_sistema, 4)
    categoria = _categoria_do_sku(db, sku, categoria_fallback)
    custo = _custo_unitario_do_sku(db, sku)
    valor_estimado = round(abs(divergencia_qtd) * custo, 2) if custo is not None else 0.0

    if divergencia_qtd == 0:
        db.add(models.MovimentacaoHistorico(
            sku=sku, almoxarifado=almoxarifado, categoria_produto=categoria,
            data_movimento=data_mov, entrada=entrada, saida=saida,
            saldo_sistema=saldo_sistema, saldo_fisico=saldo_fisico,
            divergencia=0, valor_divergencia=0, unidade=unidade,
            observacao_original=observacao, prejuizo_confirmado=prejuizo,
            status="Historico_Resolvido", lote_importacao_id=lote_id,
        ))
        # flush imediato (não deixar acumular muitos objetos pendentes pra
        # um commit em lote no final) - testado e confirmado durante a
        # preparação do deploy em nuvem: centenas de objetos pendentes
        # indo pro banco num commit só dispara um bug real de
        # desalinhamento de colunas no Postgres (recurso "insertmanyvalues"
        # do SQLAlchemy 2.x). Em SQLite isso nunca aparecia.
        db.flush()
        return "historico"

    div = models.Divergencia(
        sku=sku, almoxarifado=almoxarifado, categoria_produto=categoria,
        data_deteccao=data_mov, saldo_sistema=saldo_sistema, saldo_fisico=saldo_fisico,
        divergencia_qtd=divergencia_qtd, valor_estimado=valor_estimado, status="Aberta",
        observacao_origem=observacao, lote_importacao_id=lote_id,
    )
    db.add(div)
    db.flush()  # garante div.id antes de investigar (casos similares/self-exclusion)

    resultado_regras = investigar(db, div)
    resultado_ml = ml_predict.prever(sku, almoxarifado, categoria, divergencia_qtd, div.valor_estimado, data_mov, db=db)

    hipotese_final, confianca_final = reconciliar(
        resultado_regras["scores_normalizados"],
        resultado_ml["distribuicao"] if resultado_ml else [],
    )

    div.hipotese_regras = resultado_regras["hipotese_regras"]
    div.confianca_regras = resultado_regras["confianca_regras"]
    div.evidencias = resultado_regras["evidencias"]
    div.casos_similares = resultado_regras["casos_similares"]
    div.hipotese_ml = resultado_ml["hipotese_predita"] if resultado_ml else None
    div.confianca_ml = resultado_ml["confianca"] if resultado_ml else None
    div.distribuicao_probabilidades = resultado_ml["distribuicao"] if resultado_ml else resultado_regras["scores_normalizados"]
    div.hipotese_ia = hipotese_final
    div.confianca_ia = confianca_final

    return "divergencia"


@router.post("/custos-planilha-preco")
async def importar_custos_planilha_preco(
    arquivo: UploadFile = File(...),
    aba: str = Form("tabela de preço"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa custo unitário a partir de uma planilha de explosão de
    ficha técnica com custo por item (ex: aba 'tabela de preço' do
    arquivo de conciliação). Um mesmo SKU pode aparecer várias vezes
    nessa aba (uma vez por receita que o usa) - só atualiza o custo
    quando o valor é o mesmo em TODAS as ocorrências do SKU (evita
    gravar um custo errado quando a planilha tem valores ambíguos/
    contextuais para aquele item)."""
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")

    if aba not in wb.sheetnames:
        raise HTTPException(400, f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    ws = wb[aba]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise HTTPException(400, "Planilha vazia.")

    cabecalho = [str(c).strip() if c else "" for c in linhas[0]]
    indice = {nome: i for i, nome in enumerate(cabecalho)}
    for col_obrigatoria in ("Id_item", "custo"):
        if col_obrigatoria not in indice:
            raise HTTPException(400, f"Coluna obrigatória '{col_obrigatoria}' não encontrada. Cabeçalho: {cabecalho}")

    valores_por_sku: dict[str, set] = {}
    for linha in linhas[1:]:
        sku_bruto = linha[indice["Id_item"]]
        custo_bruto = linha[indice["custo"]]
        if sku_bruto is None or str(sku_bruto).strip() == "":
            continue
        sku = parse_sku(sku_bruto)
        custo = round(parse_decimal(custo_bruto), 6)
        valores_por_sku.setdefault(sku, set()).add(custo)

    atualizados, ambiguos, nao_encontrados = 0, [], []
    for sku, valores in valores_por_sku.items():
        if len(valores) > 1:
            ambiguos.append(sku)
            continue
        custo_unico = next(iter(valores))
        produto = db.query(models.Produto).filter_by(sku=sku).first()
        if not produto:
            nao_encontrados.append(sku)
            continue
        produto.custo_unitario = custo_unico
        atualizados += 1

    registrar_log(db, usuario.username, "importar_custos_planilha_preco",
                  detalhes={"arquivo": arquivo.filename, "atualizados": atualizados, "ambiguos": len(ambiguos), "nao_encontrados": len(nao_encontrados)})
    db.commit()
    return {
        "arquivo": arquivo.filename,
        "produtos_atualizados": atualizados,
        "skus_com_custo_ambiguo_ignorados": ambiguos,
        "skus_nao_encontrados_no_cadastro": nao_encontrados,
    }


def _ler_planilha(conteudo: bytes, aba: str) -> list[dict]:
    """Lê uma aba de Excel e devolve lista de dicts (chave = cabeçalho,
    já em minúsculo e sem espaço nas pontas, pra casar fácil com o
    de-para de cada importador)."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    if aba not in wb.sheetnames:
        raise HTTPException(400, f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    linhas_brutas = list(wb[aba].iter_rows(values_only=True))
    if not linhas_brutas:
        raise HTTPException(400, "Planilha vazia.")
    cabecalho = [str(c).strip() if c is not None else "" for c in linhas_brutas[0]]
    return [dict(zip(cabecalho, linha)) for linha in linhas_brutas[1:] if any(v is not None for v in linha)]


def _data_excel(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return parse_data(v)


@router.post("/faturamento-excel")
async def importar_faturamento_excel(
    arquivo: UploadFile = File(...), aba: str = Form("Faturamento"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db),
):
    """Faturamento é uma tabela espelhada do banco SQL da empresa - cada
    envio é o retrato atual completo, não um incremento. Por isso o
    padrão é substituir tudo, não acumular (evita duplicar toda vez que
    alguém reenvia a planilha atualizada)."""
    linhas = _ler_planilha(await arquivo.read(), aba)
    db.query(models.Faturamento).delete()
    objetos = [
        models.Faturamento(
            sku=parse_sku(l.get("SKU")), origem=l.get("Origem"),
            data_faturamento=_data_excel(l.get("Data")), quantidade=parse_decimal(l.get("Quantidade", 0)),
            descricao=l.get("Descricao"),
        )
        for l in linhas if l.get("SKU")
    ]
    db.bulk_save_objects(objetos)
    registrar_log(db, usuario.username, "importar_faturamento_excel", detalhes={"linhas": len(objetos)})
    db.commit()
    return {"arquivo": arquivo.filename, "linhas_importadas": len(objetos), "modo": "substituicao_completa"}


@router.post("/ficha-tecnica-bom-excel")
async def importar_ficha_tecnica_bom_excel(
    arquivo: UploadFile = File(...), aba: str = Form("Planilha1"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db),
):
    """Ficha técnica (BOM) rica, com subconjunto/custo/Gera_Oc - também
    substitui tudo a cada envio (mesmo motivo do faturamento)."""
    linhas = _ler_planilha(await arquivo.read(), aba)
    db.query(models.FichaTecnicaBOM).delete()
    objetos = []
    for l in linhas:
        if not l.get("Id_produto") or not l.get("Id_item"):
            continue
        objetos.append(models.FichaTecnicaBOM(
            sku_produto_final=parse_sku(l.get("Id_produto")), produto_final=l.get("Produto"),
            sku_subconjunto=parse_sku(l.get("Id_subconjunto")), subconjunto=l.get("Subconjunto"),
            sku_item=parse_sku(l.get("Id_item")), descricao_item=l.get("item"),
            qtd_padrao=parse_decimal(l.get("Qtd", 0)), unidade=l.get("item_Unidade"),
            custo=parse_decimal(l.get("custo")) if l.get("custo") is not None else None,
            tem_filho=str(l.get("Tem_Filho", "")).strip().upper() == "S",
            gera_oc=str(l.get("Gera_Oc", "")).strip().upper() == "S",
            categoria=l.get("Grupo"), linha_producao=l.get("Linha"),
        ))
    db.bulk_save_objects(objetos)
    registrar_log(db, usuario.username, "importar_ficha_tecnica_bom_excel", detalhes={"linhas": len(objetos)})
    db.commit()
    return {"arquivo": arquivo.filename, "linhas_importadas": len(objetos), "modo": "substituicao_completa"}


@router.post("/ordens-producao-excel")
async def importar_ordens_producao_excel(
    arquivo: UploadFile = File(...), aba: str = Form("Ops - PA"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db),
):
    linhas = _ler_planilha(await arquivo.read(), aba)
    db.query(models.OrdemProducao).delete()
    objetos = [
        models.OrdemProducao(
            numero_op=str(l.get("Número da OP")).strip(), sku_produto_final=parse_sku(l.get("SKU Produto Final")),
            descricao_produto=l.get("Desc_Prod"),
            data_registro=_data_excel(l.get("Dt_Registro")), data_producao=_data_excel(l.get("Dt_Producao")),
            status=l.get("Status"),
            qtd_prevista=parse_decimal(l.get("Qtd_Prevista", 0)), qtd_produzida=parse_decimal(l.get("Qtd_Prod", 0)),
            qtd_saldo=parse_decimal(l.get("Qtd_Saldo", 0)),
        )
        for l in linhas if l.get("Número da OP") is not None
    ]
    db.bulk_save_objects(objetos)
    registrar_log(db, usuario.username, "importar_ordens_producao_excel", detalhes={"linhas": len(objetos)})
    db.commit()
    return {"arquivo": arquivo.filename, "linhas_importadas": len(objetos), "modo": "substituicao_completa"}


@router.post("/consumo-op-excel")
async def importar_consumo_op_excel(
    arquivo: UploadFile = File(...), aba: str = Form("OPs - Mat.P"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db),
):
    linhas = _ler_planilha(await arquivo.read(), aba)
    db.query(models.ConsumoOP).delete()
    objetos, ignorados = [], 0
    for l in linhas:
        numero_op = l.get("Número da OP")
        if numero_op is None:
            ignorados += 1
            continue
        objetos.append(models.ConsumoOP(
            numero_op=str(numero_op).strip(), sku_produto_final=parse_sku(l.get("SKU Produto Final")),
            sku_material=parse_sku(l.get("Material")), descricao_material=l.get("Desc_Material"),
            qtd_consumo=parse_decimal(l.get("qtd_consumo", 0)), qtd_previsto=parse_decimal(l.get("Qtd_Previsto", 0)),
            qtd_diferenca=parse_decimal(l.get("Qtd_Dif", 0)),
            data_registro=_data_excel(l.get("Dt_Registro")), data_producao=_data_excel(l.get("Dt_Producao")),
            status=l.get("Status"),
        ))
    db.bulk_save_objects(objetos)
    registrar_log(db, usuario.username, "importar_consumo_op_excel", detalhes={"linhas": len(objetos), "ignorados": ignorados})
    db.commit()
    return {"arquivo": arquivo.filename, "linhas_importadas": len(objetos), "ignorados_sem_numero_op": ignorados, "modo": "substituicao_completa"}


@router.post("/transferencias-excel")
async def importar_transferencias_excel(
    arquivo: UploadFile = File(...), aba: str = Form("Transferencia"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db),
):
    linhas = _ler_planilha(await arquivo.read(), aba)
    db.query(models.Transferencia).delete()
    objetos = [
        models.Transferencia(
            sku=parse_sku(l.get("SKU")), descricao=l.get("descricao"),
            data_saida=_data_excel(l.get("Data de Saída")), data_entrada=_data_excel(l.get("Data de Entrada")),
            documento=str(l.get("doc")) if l.get("doc") is not None else None,
            almoxarifado_origem=normalizar_almoxarifado(l.get("Almoxarifado de Origem")) if l.get("Almoxarifado de Origem") else None,
            almoxarifado_destino=normalizar_almoxarifado(l.get("Almoxarifado de Destino")) if l.get("Almoxarifado de Destino") else None,
            quantidade=parse_decimal(l.get("Quantidade", 0)), lote=l.get("lote"),
        )
        for l in linhas if l.get("SKU")
    ]
    db.bulk_save_objects(objetos)
    registrar_log(db, usuario.username, "importar_transferencias_excel", detalhes={"linhas": len(objetos)})
    db.commit()
    return {"arquivo": arquivo.filename, "linhas_importadas": len(objetos), "modo": "substituicao_completa"}


@router.post("/custos")
async def importar_custos(
    arquivo: UploadFile = File(...),
    aba: str | None = Form(None),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Duas colunas: sku, custo_unitario. Aceita CSV ou Excel (.xlsx/.xlsm)
    - detecta pela extensão do arquivo. Atualiza o cadastro de produtos -
    SKUs que ainda não existem no cadastro são ignorados (a lista de
    produtos é gerenciada pela importação de produtos, não por aqui) e
    ficam relatados em erros/ignorados para você revisar."""
    conteudo = await arquivo.read()
    nome = (arquivo.filename or "").lower()

    linhas: list[dict] = []
    if nome.endswith((".xlsx", ".xlsm")):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
        except Exception as e:
            raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")
        nome_aba = aba or wb.sheetnames[0]
        if nome_aba not in wb.sheetnames:
            raise HTTPException(400, f"Aba '{nome_aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        ws = wb[nome_aba]
        linhas_brutas = list(ws.iter_rows(values_only=True))
        if not linhas_brutas:
            raise HTTPException(400, "Planilha vazia.")
        cabecalho = [str(c).strip().lower() if c else "" for c in linhas_brutas[0]]
        if "sku" not in cabecalho or "custo_unitario" not in cabecalho:
            raise HTTPException(400, f"Colunas 'sku' e 'custo_unitario' não encontradas. Cabeçalho: {cabecalho}")
        idx_sku, idx_custo = cabecalho.index("sku"), cabecalho.index("custo_unitario")
        for linha in linhas_brutas[1:]:
            linhas.append({"sku": linha[idx_sku], "custo_unitario": linha[idx_custo]})
    else:
        try:
            texto = conteudo.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(400, "Não consegui ler o arquivo como CSV (texto). Se for um Excel, salve como .xlsx ou .csv antes de enviar.")
        leitor = csv_module.DictReader(io.StringIO(texto))
        linhas = list(leitor)

    atualizados, ignorados, erros = 0, [], []
    for i, row in enumerate(linhas, start=1):
        try:
            sku = parse_sku(row["sku"])
            if sku is None:
                continue
            custo = parse_decimal(row["custo_unitario"])
            produto = db.query(models.Produto).filter_by(sku=sku).first()
            if not produto:
                ignorados.append(sku)
                continue
            produto.custo_unitario = custo
            atualizados += 1
        except Exception as e:
            erros.append(f"linha {i}: {e}")
    registrar_log(db, usuario.username, "importar_custos", detalhes={"arquivo": arquivo.filename, "atualizados": atualizados})
    db.commit()
    return {
        "arquivo": arquivo.filename,
        "produtos_atualizados": atualizados,
        "skus_nao_encontrados_no_cadastro": ignorados,
        "erros": erros,
    }


@router.get("/almoxarifados")
def listar_almoxarifados(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    return [{"codigo": a.codigo, "nome": a.nome_exibicao} for a in db.query(models.Almoxarifado).filter_by(ativo=True).all()]


@router.get("/lotes")
def listar_lotes(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Lista as últimas importações de movimentação (CSV/Excel), pra dar
    pra ver o que foi importado e, se precisar, excluir uma inteira."""
    lotes = db.query(models.LoteImportacao).order_by(models.LoteImportacao.criado_em.desc()).limit(50).all()
    return [
        {
            "id": l.id, "tipo": l.tipo, "arquivo_origem": l.arquivo_origem, "almoxarifado": l.almoxarifado,
            "usuario": l.usuario, "criado_em": l.criado_em.isoformat(), "linhas_processadas": l.linhas_processadas,
            "divergencias_criadas": l.divergencias_criadas, "historico_criado": l.historico_criado,
        }
        for l in lotes
    ]


@router.delete("/lotes/{lote_id}")
def excluir_lote(lote_id: int, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Desfaz uma importação inteira: remove as divergências e o
    histórico criados por ela (e o feedback de ML ligado a essas
    divergências, pra não deixar lixo no próximo retreino)."""
    lote = db.query(models.LoteImportacao).get(lote_id)
    if not lote:
        raise HTTPException(404, "Lote de importação não encontrado.")

    divergencias = db.query(models.Divergencia).filter_by(lote_importacao_id=lote_id).all()
    ids_divergencias = [d.id for d in divergencias]
    if ids_divergencias:
        db.query(models.CasoMLFeedback).filter(models.CasoMLFeedback.divergencia_id.in_(ids_divergencias)).delete(synchronize_session=False)
    qtd_div = db.query(models.Divergencia).filter_by(lote_importacao_id=lote_id).delete(synchronize_session=False)
    qtd_hist = db.query(models.MovimentacaoHistorico).filter_by(lote_importacao_id=lote_id).delete(synchronize_session=False)

    registrar_log(db, usuario.username, "excluir_lote_importacao", entidade="lote_importacao", entidade_id=lote_id,
                  detalhes={"arquivo": lote.arquivo_origem, "divergencias_removidas": qtd_div, "historico_removido": qtd_hist})
    db.delete(lote)
    db.commit()
    return {"ok": True, "divergencias_removidas": qtd_div, "historico_removido": qtd_hist}


@router.post("/movimentacao")
async def importar_movimentacao_csv(
    arquivo: UploadFile = File(...),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importação via CSV - cada linha já traz sua própria coluna de
    almoxarifado. Colunas esperadas: sku, almoxarifado, data_movimento,
    entrada, saida, saldo_sistema, saldo_fisico, unidade."""
    lote = models.LoteImportacao(tipo="movimentacao_csv", arquivo_origem=arquivo.filename, usuario=usuario.username)
    db.add(lote)
    db.flush()

    conteudo = (await arquivo.read()).decode("utf-8-sig")
    leitor = csv_module.DictReader(io.StringIO(conteudo))
    total, hist, div, erros = 0, 0, 0, []
    for i, row in enumerate(leitor, start=1):
        total += 1
        try:
            destino = processar_linha_movimentacao(db, row, lote_id=lote.id)
            if destino == "historico":
                hist += 1
            else:
                div += 1
        except Exception as e:
            erros.append(f"linha {i}: {e}")

    lote.linhas_processadas, lote.divergencias_criadas, lote.historico_criado = total, div, hist
    registrar_log(db, usuario.username, "importar_movimentacao_csv", entidade="lote_importacao", entidade_id=lote.id,
                  detalhes={"arquivo": arquivo.filename, "linhas": total, "divergencias": div})
    db.commit()
    return {
        "lote_id": lote.id,
        "arquivo": arquivo.filename,
        "linhas_processadas": total,
        "inseridas_historico": hist,
        "inseridas_divergencias": div,
        "erros": erros,
    }


@router.post("/movimentacao-excel")
async def importar_movimentacao_excel(
    arquivo: UploadFile = File(...),
    almoxarifado: str = Form(...),
    aba: str | None = Form(None),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importação direta do arquivo real da operação (.xlsx/.xlsm), sem
    precisar converter pra CSV. O arquivo não traz almoxarifado por linha
    (cada planilha já é de um almoxarifado só), então ele é informado no
    formulário. Espera a aba 'Movimentação' com as colunas: Grupo, Data,
    Id_Produto, Unid., Sistema, Entrada, Saídas, Contagem, Obs, Prejuízo."""
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")

    nome_aba = aba or ("Movimentação" if "Movimentação" in wb.sheetnames else wb.sheetnames[0])
    if nome_aba not in wb.sheetnames:
        raise HTTPException(400, f"Aba '{nome_aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    ws = wb[nome_aba]

    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise HTTPException(400, "Planilha vazia.")
    cabecalho = [str(c).strip() if c else "" for c in linhas[0]]
    indice = {nome: i for i, nome in enumerate(cabecalho)}

    faltando = [nome_col for nome_col in COLUNAS_EXCEL.values() if nome_col not in indice]
    if faltando:
        raise HTTPException(400, f"Colunas esperadas não encontradas na aba '{nome_aba}': {faltando}. Cabeçalho encontrado: {cabecalho}")

    lote = models.LoteImportacao(tipo="movimentacao_excel", arquivo_origem=arquivo.filename, almoxarifado=almoxarifado, usuario=usuario.username)
    db.add(lote)
    db.flush()

    total, hist, div, erros = 0, 0, 0, []
    for i, linha in enumerate(linhas[1:], start=2):
        sku_bruto = linha[indice[COLUNAS_EXCEL["sku"]]]
        if sku_bruto is None or str(sku_bruto).strip() == "":
            continue  # linha em branco / separadora
        total += 1
        try:
            data_bruta = linha[indice[COLUNAS_EXCEL["data"]]]
            data_mov = data_bruta.date() if isinstance(data_bruta, datetime) else parse_data(data_bruta)
            obs = linha[indice[COLUNAS_EXCEL["obs"]]]
            prejuizo_bruto = linha[indice[COLUNAS_EXCEL["prejuizo"]]]
            row = {
                "sku": sku_bruto,
                "almoxarifado": almoxarifado,
                "data_movimento": data_mov,
                "entrada": linha[indice[COLUNAS_EXCEL["entrada"]]],
                "saida": linha[indice[COLUNAS_EXCEL["saida"]]],
                "saldo_sistema": linha[indice[COLUNAS_EXCEL["saldo_sistema"]]],
                "saldo_fisico": linha[indice[COLUNAS_EXCEL["saldo_fisico"]]],
                "unidade": linha[indice[COLUNAS_EXCEL["unidade"]]],
            }
            categoria_fallback = linha[indice[COLUNAS_EXCEL["grupo"]]]
            destino = processar_linha_movimentacao(
                db, row, almoxarifado_forcado=almoxarifado, categoria_fallback=categoria_fallback,
                observacao=obs, prejuizo=str(prejuizo_bruto).strip().lower() == "sim", lote_id=lote.id,
            )
            if destino == "historico":
                hist += 1
            else:
                div += 1
        except Exception as e:
            erros.append(f"linha {i}: {e}")

    lote.linhas_processadas, lote.divergencias_criadas, lote.historico_criado = total, div, hist
    registrar_log(db, usuario.username, "importar_movimentacao_excel", entidade="lote_importacao", entidade_id=lote.id,
                  detalhes={"arquivo": arquivo.filename, "almoxarifado": almoxarifado, "linhas": total, "divergencias": div})
    db.commit()
    return {
        "lote_id": lote.id,
        "arquivo": arquivo.filename,
        "aba": nome_aba,
        "almoxarifado": almoxarifado,
        "linhas_processadas": total,
        "inseridas_historico": hist,
        "inseridas_divergencias": div,
        "erros": erros,
    }
