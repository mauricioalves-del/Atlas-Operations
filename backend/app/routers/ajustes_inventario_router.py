"""
Importador e leitura da tabela OFICIAL de ajustes de inventário ("Ace4" -
aba "Estoque" da planilha "Inventários"). Diferente do fechamento bruto
(fechamento_router.py / ItemFechamento), que registra TODO item sinalizado
como divergente no momento da contagem - inclusive os que a operação
nunca chegou a ajustar de fato, por problema de processo -, esta tabela
só tem o que foi de fato conciliado e processado. É a fonte usada pelo
painel "Fluxo de Inventário" do Mapeamento de Passivos (ver
baixas_operacionais_router.py).

Contexto (explicado pelo Maurício): até o meio de 2026, TODA baixa
operacional (passivo) passava pelo módulo de inventário, misturada com os
ajustes de inventário de fato. A coluna "Inventário" da planilha separa
os dois: "Sim" = ajuste de inventário real; "Não" = baixa de passivo que
só passou por ali, mas já é contabilizada em outro lugar (BaixaOperacional
- ver dashboard de Mapeamento de Passivos) - por isso é DESCONSIDERADA
aqui, pra não contar duas vezes. Linhas de antes dessa separação existir
(Jul-Dez/2025) vêm com o ANO na coluna em vez de Sim/Não - por decisão do
Maurício, essas ficam de fora do cálculo de Entradas/Saídas (não tem como
saber com certeza se eram ajuste ou passivo).

A partir de PADRONIZACAO_NOTA_FISCAL_DESDE (jul/2026), as baixas de
passivo passaram a vir só por nota fiscal (fora dessa planilha) - então
qualquer lançamento de inventário a partir daí já é ajuste de inventário
automaticamente, mesmo sem a coluna "Inventário" preenchida ou marcada
"Não" (Maurício confirmou: entrada e saída contam igual)."""
import io
from datetime import date, datetime
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query
from sqlalchemy.orm import Session
import openpyxl

from .. import models, schemas
from ..database import get_db
from ..csv_utils import parse_sku, parse_data, parse_decimal, limpar_texto
from ..hipoteses_config import normalizar_almoxarifado
from ..deps import requer_papel, obter_usuario_atual

router = APIRouter(prefix="/ajustes-inventario", tags=["ajustes_inventario_oficial"])

ABA_PADRAO = "Estoque"

# De-para das colunas da planilha oficial (aba "Estoque" de Inventários.xlsx)
# pros campos internos - ajuste aqui se a planilha da empresa usar nomes
# um pouco diferentes.
COLUNAS = {
    "sku": "Id_Produto",
    "status": "Status",
    "id_invent": "Id_Invent",
    "data": "Dt_Invent",
    "almoxarifado": "Almox",
    "descricao": "Descricao",
    "id_lote": "Id_Lote",
    "qtd_sistema": "Qtd",
    "qtd_contagem": "Cont1",
    "ajuste": "Ajuste",
    "custo": "Custo",
    "valor_total": "Vlr_Total",
    "grupo": "Grupo",
    "obs": "Obs",
    "inventario_flag": "Inventário",
}

# A partir desta data, todo lançamento de inventário já é ajuste de
# inventário automaticamente (baixas de passivo passaram a vir só por nota
# fiscal, fora desta planilha) - confirmado pelo Maurício, vale pra entrada
# E saída igualmente.
PADRONIZACAO_NOTA_FISCAL_DESDE = date(2026, 7, 1)


def conta_como_ajuste_inventario(dt_invent, inventario_flag_bruto) -> bool:
    """Regra confirmada com o Maurício (não muda sem confirmar de novo):
    - a partir de jul/2026: SEMPRE conta como ajuste, não importa a coluna;
    - antes disso: só conta quando a coluna for exatamente "Sim" - "Não" e
      qualquer outra coisa (inclusive o ano legado "2025"/"2026" de antes
      da separação existir) NÃO conta."""
    if dt_invent and dt_invent >= PADRONIZACAO_NOTA_FISCAL_DESDE:
        return True
    flag = str(inventario_flag_bruto).strip().lower() if inventario_flag_bruto is not None else ""
    return flag == "sim"


@router.post("/importar")
async def importar_ajustes_inventario(
    arquivo: UploadFile = File(...),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa a planilha oficial de ajustes de inventário (aba "Estoque").
    Cada linha é um ajuste já conciliado - não faz upsert por chave (a
    mesma combinação Id_Produto+Id_Invent+Id_Lote pode legitimamente
    aparecer mais de uma vez na planilha, com ajustes diferentes - são
    lançamentos distintos, não duplicata), então re-importar o mesmo
    arquivo duas vezes cria linhas repetidas. Se um Id_Invent já importado
    antes aparecer de novo no arquivo, isso é avisado na resposta (não
    bloqueia), pra dar visibilidade sem travar uma reconciliação legítima
    que reprocessa um inventário antigo."""
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")

    aba = ABA_PADRAO if ABA_PADRAO in wb.sheetnames else None
    if aba is None:
        # fallback: procura a primeira aba cujo cabeçalho bate com o esperado
        for nome in wb.sheetnames:
            primeira_linha = next(wb[nome].iter_rows(min_row=1, max_row=1, values_only=True), None)
            cabecalho_candidato = {str(c).strip() if c else "" for c in (primeira_linha or [])}
            if COLUNAS["sku"] in cabecalho_candidato and COLUNAS["ajuste"] in cabecalho_candidato:
                aba = nome
                break
    if aba is None:
        raise HTTPException(400, f"Não encontrei a aba '{ABA_PADRAO}' (nem nenhuma outra com as colunas esperadas). Abas disponíveis: {wb.sheetnames}")

    ws = wb[aba]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise HTTPException(400, "Planilha vazia.")

    cabecalho = [str(c).strip() if c else "" for c in linhas[0]]
    indice = {nome: i for i, nome in enumerate(cabecalho)}
    faltando = [c for c in [COLUNAS["sku"], COLUNAS["data"], COLUNAS["ajuste"]] if c not in indice]
    if faltando:
        raise HTTPException(400, f"Colunas esperadas não encontradas: {faltando}. Cabeçalho: {cabecalho}")

    def val(linha, chave):
        col = COLUNAS[chave]
        return linha[indice[col]] if col in indice else None

    ids_invent_ja_existentes = {r[0] for r in db.query(models.AjusteInventarioOficial.id_invent).distinct().all() if r[0] is not None}
    ids_invent_repetidos = set()

    lote = models.LoteAjusteInventario(arquivo_origem=arquivo.filename, aba_usada=aba, criado_por=usuario.username)
    db.add(lote)
    db.flush()  # garante lote.id antes do loop, pra já linkar cada linha

    total, importadas, contadas_ajuste, ignoradas_nao, ignoradas_legado, erros = 0, 0, 0, 0, 0, []
    valor_total_ajuste = 0.0
    # Alguns exports do sistema de origem repetem a MESMA linha inteira (visto
    # em planilhas reais: dezenas de linhas 100% idênticas, não é "mesma
    # chave com ajuste diferente" - é a linha inteira duplicada por engano
    # do export). Sem isso, cada duplicata dobra a quantidade/valor daquele
    # ajuste. Só pula duplicata EXATA (todas as colunas físicas iguais)
    # dentro do MESMO arquivo - não afeta o caso legítimo de duas linhas com
    # a mesma chave natural mas valores diferentes, nem reimportações
    # intencionais em arquivos/lotes diferentes.
    linhas_vistas_no_arquivo = set()
    duplicadas_no_arquivo = 0

    for i, linha in enumerate(linhas[1:], start=2):
        sku_bruto = val(linha, "sku")
        if sku_bruto is None or str(sku_bruto).strip() == "":
            continue
        total += 1
        try:
            sku = parse_sku(sku_bruto)
            data_bruta = val(linha, "data")
            dt_invent = data_bruta.date() if isinstance(data_bruta, datetime) else parse_data(data_bruta)
            almoxarifado_origem = val(linha, "almoxarifado")
            almoxarifado = normalizar_almoxarifado(almoxarifado_origem)
            id_invent_bruto = val(linha, "id_invent")
            id_invent = int(id_invent_bruto) if id_invent_bruto is not None and str(id_invent_bruto).strip() != "" else None

            status_bruto = limpar_texto(val(linha, "status"))
            id_lote_bruto = limpar_texto(val(linha, "id_lote"))
            qtd_sistema = parse_decimal(val(linha, "qtd_sistema"))
            qtd_contagem = parse_decimal(val(linha, "qtd_contagem"))
            ajuste_qtd = parse_decimal(val(linha, "ajuste"))
            custo_unitario = parse_decimal(val(linha, "custo"))
            valor_total_linha = parse_decimal(val(linha, "valor_total"))

            chave_linha = (
                sku, status_bruto, id_invent, dt_invent, limpar_texto(almoxarifado_origem), id_lote_bruto,
                qtd_sistema, qtd_contagem, ajuste_qtd, custo_unitario, valor_total_linha,
            )
            if chave_linha in linhas_vistas_no_arquivo:
                duplicadas_no_arquivo += 1
                continue
            linhas_vistas_no_arquivo.add(chave_linha)

            if id_invent is not None and id_invent in ids_invent_ja_existentes:
                ids_invent_repetidos.add(id_invent)

            inventario_flag_bruto = val(linha, "inventario_flag")
            conta = conta_como_ajuste_inventario(dt_invent, inventario_flag_bruto)

            registro = models.AjusteInventarioOficial(
                lote_id=lote.id,
                sku=sku, status=status_bruto, id_invent=id_invent, dt_invent=dt_invent,
                almoxarifado=almoxarifado, almoxarifado_origem=limpar_texto(almoxarifado_origem),
                descricao_produto=limpar_texto(val(linha, "descricao")), id_lote=id_lote_bruto,
                qtd_sistema=qtd_sistema, qtd_contagem=qtd_contagem,
                ajuste_qtd=ajuste_qtd, custo_unitario=custo_unitario,
                valor_total=valor_total_linha, categoria_produto=limpar_texto(val(linha, "grupo")),
                observacao=limpar_texto(val(linha, "obs")),
                inventario_flag_bruto=limpar_texto(str(inventario_flag_bruto)) if inventario_flag_bruto is not None else None,
                conta_como_ajuste_inventario=conta,
                arquivo_origem=arquivo.filename, criado_por=usuario.username,
            )
            db.add(registro)
            importadas += 1
            flag_lower = str(inventario_flag_bruto).strip().lower() if inventario_flag_bruto is not None else ""
            if conta:
                contadas_ajuste += 1
                valor_total_ajuste += valor_total_linha
            elif flag_lower == "não":
                ignoradas_nao += 1
            else:
                ignoradas_legado += 1  # ano legado (pré-separação Sim/Não) ou vazio
        except Exception as e:
            erros.append(f"Linha {i} (SKU {sku_bruto}): {e}")

    lote.total_linhas = total
    lote.importadas = importadas
    lote.contadas_como_ajuste_inventario = contadas_ajuste
    lote.ignoradas_flag_nao = ignoradas_nao
    lote.ignoradas_legado_pre_separacao = ignoradas_legado
    lote.valor_total_ajustes_contados = round(valor_total_ajuste, 2)
    db.commit()
    return {
        "lote_id": lote.id,
        "arquivo": arquivo.filename, "aba_usada": aba,
        "total_linhas": total, "importadas": importadas,
        "contadas_como_ajuste_inventario": contadas_ajuste,
        "ignoradas_flag_nao": ignoradas_nao,
        "ignoradas_legado_pre_separacao": ignoradas_legado,
        "valor_total_ajustes_contados": round(valor_total_ajuste, 2),
        "ids_invent_repetidos": sorted(ids_invent_repetidos),
        "duplicadas_no_arquivo": duplicadas_no_arquivo,
        "erros": erros,
    }


@router.get("/lotes")
def listar_lotes_ajuste_inventario(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Histórico de importações (lotes) da planilha oficial de ajustes de
    inventário, mais recente primeiro - alimenta a lista da tela Importar
    (com opção de excluir um lote errado/duplicado)."""
    lotes = db.query(models.LoteAjusteInventario).order_by(models.LoteAjusteInventario.id.desc()).all()
    return [
        {
            "id": l.id,
            "arquivo_origem": l.arquivo_origem,
            "aba_usada": l.aba_usada,
            "criado_por": l.criado_por,
            "criado_em": l.criado_em.isoformat() if l.criado_em else None,
            "total_linhas": l.total_linhas,
            "importadas": l.importadas,
            "contadas_como_ajuste_inventario": l.contadas_como_ajuste_inventario,
            "ignoradas_flag_nao": l.ignoradas_flag_nao,
            "ignoradas_legado_pre_separacao": l.ignoradas_legado_pre_separacao,
            "valor_total_ajustes_contados": l.valor_total_ajustes_contados,
        }
        for l in lotes
    ]


@router.delete("/lotes/{lote_id}")
def excluir_lote_ajuste_inventario(
    lote_id: int,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Exclui um lote (importação) e todas as linhas de
    AjusteInventarioOficial vinculadas a ele - pra corrigir upload
    duplicado ou feito errado, "cada importação é uma coisa" que pode ser
    desfeita sem afetar as outras."""
    lote = db.query(models.LoteAjusteInventario).get(lote_id)
    if lote is None:
        raise HTTPException(404, "Lote não encontrado.")
    linhas_removidas = db.query(models.AjusteInventarioOficial).filter(models.AjusteInventarioOficial.lote_id == lote_id).delete(synchronize_session=False)
    db.delete(lote)
    db.commit()
    return {"lote_id": lote_id, "linhas_removidas": linhas_removidas}


# ---------------------------------------------------------------------------
# Justificativas de ajuste de inventário - por que um ajuste específico
# aconteceu e qual solução foi aplicada. Espelha o padrão de
# AcaoPosInventario (fechamento_router.py: listar/detalhar/criar/atualizar/
# excluir), aberto a partir de uma linha do painel "Fluxo de Inventário" do
# Mapeamento de Passivos (ver baixas_operacionais_router.py e app.js).
# ---------------------------------------------------------------------------

@router.get("/justificativas", response_model=list[schemas.JustificativaAjusteInventarioOut])
def listar_justificativas(
    status: str | None = None,
    ajuste_id: int | None = None,
    sku: str | None = None,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Lista justificativas de ajuste de inventário - alimenta a tabela
    "Justificativas de Ajustes de Inventário" no Mapeamento de Passivos, e
    também é usada pelo pop de justificativa pra achar se já existe uma
    justificativa em aberto pra aquele ajuste (filtro por ajuste_id)."""
    q = db.query(models.JustificativaAjusteInventario)
    if status:
        q = q.filter(models.JustificativaAjusteInventario.status == status)
    if ajuste_id:
        q = q.filter(models.JustificativaAjusteInventario.ajuste_id == ajuste_id)
    if sku:
        q = q.filter(models.JustificativaAjusteInventario.sku == sku)
    return q.order_by(models.JustificativaAjusteInventario.criado_em.desc()).all()


@router.get("/justificativas/{justificativa_id}", response_model=schemas.JustificativaAjusteInventarioOut)
def detalhar_justificativa(justificativa_id: int, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    justificativa = db.query(models.JustificativaAjusteInventario).get(justificativa_id)
    if not justificativa:
        raise HTTPException(404, "Justificativa não encontrada.")
    return justificativa


@router.post("/justificativas", response_model=schemas.JustificativaAjusteInventarioOut)
def criar_justificativa(
    payload: schemas.JustificativaAjusteInventarioCreate,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    nova = models.JustificativaAjusteInventario(**payload.model_dump(), status="Pendente", criado_por=usuario.username)
    db.add(nova)
    db.commit()
    db.refresh(nova)
    return nova


@router.patch("/justificativas/{justificativa_id}", response_model=schemas.JustificativaAjusteInventarioOut)
def atualizar_justificativa(
    justificativa_id: int,
    payload: schemas.JustificativaAjusteInventarioAtualizar,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    justificativa = db.query(models.JustificativaAjusteInventario).get(justificativa_id)
    if not justificativa:
        raise HTTPException(404, "Justificativa não encontrada.")
    dados = payload.model_dump(exclude_unset=True)
    for campo, valor in dados.items():
        setattr(justificativa, campo, valor)
    if dados.get("status") == "Concluida" and not justificativa.concluido_em:
        justificativa.concluido_em = datetime.utcnow()
    db.commit()
    db.refresh(justificativa)
    return justificativa


@router.delete("/justificativas/{justificativa_id}")
def excluir_justificativa(justificativa_id: int, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    justificativa = db.query(models.JustificativaAjusteInventario).get(justificativa_id)
    if not justificativa:
        raise HTTPException(404, "Justificativa não encontrada.")
    db.delete(justificativa)
    db.commit()
    return {"ok": True}
