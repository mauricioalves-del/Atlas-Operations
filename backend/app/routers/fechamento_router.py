import io
from datetime import datetime
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.orm import Session
import openpyxl

from .. import models, schemas
from ..database import get_db
from ..csv_utils import parse_sku, parse_data, parse_decimal
from ..hipoteses_config import normalizar_almoxarifado, buscar_evidencias_texto
from ..investigation import investigar, reconciliar
from ..ml import predict as ml_predict
from ..deps import requer_papel, obter_usuario_atual
from ..audit import registrar_log

router = APIRouter(prefix="/fechamentos", tags=["fechamento_inventario"])

ABA_PADRAO = "Saldo de estoque - ace4"

# De-para das colunas da planilha de fechamento (conciliação contábil x
# físico) para os campos internos. Ajuste aqui se a planilha da sua
# empresa usar nomes um pouco diferentes.
COLUNAS = {
    "grupo": "Grupo",
    "sku": "Produto",
    "descricao": "Descrição",
    "almoxarifado": "Almoxarifado",
    "unidade": "Unid.",
    "qtd_sistema": "Qtd.",
    "qtd_contagem": "Contagem 1",
    "status": "Status",
    "resumo": "Resumo",
    "valor": "Vl",
    "percentual": "%",
    "obs": "Obs",
    "obs_pos_inventario": "obs pós inv",
    "obs2": "obs2",
    "data": "Data",
}


def _custo_unitario_do_sku(db: Session, sku: str):
    p = db.query(models.Produto).filter_by(sku=sku).first()
    return p.custo_unitario if p else None


def _categoria_do_sku(db: Session, sku: str, fallback=None):
    p = db.query(models.Produto).filter_by(sku=sku).first()
    if p and p.categoria_produto:
        return p.categoria_produto
    return fallback or "Desconhecido"


@router.post("/importar")
async def importar_fechamento(
    arquivo: UploadFile = File(...),
    aba: str = Form(ABA_PADRAO),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa uma planilha de fechamento de inventário (conciliação
    contábil x físico). Cada linha divergente é investigada pelo mesmo
    motor (regras + ML + texto) usado no resto do sistema, e comparada
    com fechamentos anteriores do mesmo SKU+almoxarifado para calcular
    recorrência - itens que já divergiram antes ganham destaque (⭐)."""
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")

    if aba not in wb.sheetnames:
        raise HTTPException(400, f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    ws = wb[aba]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        raise HTTPException(400, "Planilha vazia.")

    cabecalho = [str(c).strip() if c else "" for c in linhas[0]]
    indice = {nome: i for i, nome in enumerate(cabecalho)}
    faltando = [c for c in [COLUNAS["sku"], COLUNAS["almoxarifado"], COLUNAS["status"], COLUNAS["data"]] if c not in indice]
    if faltando:
        raise HTTPException(400, f"Colunas esperadas não encontradas: {faltando}. Cabeçalho: {cabecalho}")

    def val(linha, chave):
        col = COLUNAS[chave]
        return linha[indice[col]] if col in indice else None

    # descobre almoxarifado/data predominantes para o registro de fechamento
    primeira_data, primeiro_almox = None, None
    fechamento = models.FechamentoInventario(arquivo_origem=arquivo.filename, criado_por=usuario.username)
    db.add(fechamento)
    db.flush()

    total, divergentes, valor_total, erros = 0, 0, 0.0, []

    for i, linha in enumerate(linhas[1:], start=2):
        sku_bruto = val(linha, "sku")
        if sku_bruto is None or str(sku_bruto).strip() == "":
            continue
        try:
            sku = parse_sku(sku_bruto)
            almoxarifado = normalizar_almoxarifado(val(linha, "almoxarifado"))
            data_bruta = val(linha, "data")
            data_fechamento = data_bruta.date() if isinstance(data_bruta, datetime) else parse_data(data_bruta)
            if primeira_data is None:
                primeira_data, primeiro_almox = data_fechamento, almoxarifado
                fechamento.data_fechamento = data_fechamento
                fechamento.almoxarifado = almoxarifado

            qtd_sistema = parse_decimal(val(linha, "qtd_sistema"))
            qtd_contagem = parse_decimal(val(linha, "qtd_contagem"))
            status_txt = str(val(linha, "status") or "").strip().lower()
            divergente = not status_txt.startswith("sem")

            divergencia_qtd = round(qtd_contagem - qtd_sistema, 4) if divergente else 0.0
            valor_planilha = parse_decimal(val(linha, "valor"))
            if divergente:
                custo = _custo_unitario_do_sku(db, sku)
                valor_estimado = valor_planilha if valor_planilha else (round(abs(divergencia_qtd) * custo, 2) if custo else 0.0)
            else:
                valor_estimado = 0.0

            categoria = _categoria_do_sku(db, sku, val(linha, "grupo"))
            observacao = val(linha, "obs_pos_inventario") or val(linha, "obs") or val(linha, "resumo")

            percentual_planilha = val(linha, "percentual")
            if percentual_planilha is not None:
                percentual_acuracia = float(percentual_planilha)
            elif qtd_sistema:
                percentual_acuracia = round(min(qtd_contagem, qtd_sistema) / qtd_sistema, 4)
            else:
                percentual_acuracia = 1.0 if not divergente else None

            # --- recorrência: já divergiu em fechamento anterior deste sku+almoxarifado? ---
            recorrencias = (
                db.query(models.ItemFechamento)
                .filter(
                    models.ItemFechamento.sku == sku,
                    models.ItemFechamento.almoxarifado == almoxarifado,
                    models.ItemFechamento.divergente.is_(True),
                    models.ItemFechamento.fechamento_id != fechamento.id,
                )
                .count()
            )

            divergencia_id = None
            movimentacao_historico_id = None
            if divergente:
                div = models.Divergencia(
                    sku=sku, almoxarifado=almoxarifado, categoria_produto=categoria,
                    data_deteccao=data_fechamento, saldo_sistema=qtd_sistema, saldo_fisico=qtd_contagem,
                    divergencia_qtd=divergencia_qtd, valor_estimado=valor_estimado, status="Aberta",
                    observacao_origem=observacao, origem="fechamento_inventario",
                )
                db.add(div)
                db.flush()

                resultado_regras = investigar(db, div)
                resultado_ml = ml_predict.prever(sku, almoxarifado, categoria, divergencia_qtd, valor_estimado, data_fechamento)
                hipotese_final, confianca_final = reconciliar(
                    resultado_regras["scores_normalizados"], resultado_ml["distribuicao"] if resultado_ml else []
                )
                div.hipotese_regras = resultado_regras["hipotese_regras"]
                div.confianca_regras = resultado_regras["confianca_regras"]
                div.evidencias = resultado_regras["evidencias"]
                div.casos_similares = resultado_regras["casos_similares"]
                div.hipotese_ml = resultado_ml["hipotese_predita"] if resultado_ml else None
                div.confianca_ml = resultado_ml["confianca"] if resultado_ml else None
                div.distribuicao_probabilidades = resultado_ml["distribuicao"] if resultado_ml else resultado_regras["scores_normalizados"]
                div.hipotese_ia = hipotese_final
                div.confianca_ia = confianca_final
                divergencia_id = div.id
                divergentes += 1
                valor_total += abs(valor_estimado)
            else:
                hist = models.MovimentacaoHistorico(
                    sku=sku, almoxarifado=almoxarifado, categoria_produto=categoria,
                    data_movimento=data_fechamento, entrada=0, saida=0,
                    saldo_sistema=qtd_sistema, saldo_fisico=qtd_contagem,
                    divergencia=0, valor_divergencia=0, unidade=val(linha, "unidade"),
                    observacao_original=observacao, status="Historico_Resolvido",
                    origem="fechamento_inventario",
                )
                db.add(hist)
                db.flush()
                movimentacao_historico_id = hist.id

            novo_item = models.ItemFechamento(
                fechamento_id=fechamento.id, sku=sku, descricao_produto=val(linha, "descricao"),
                almoxarifado=almoxarifado, categoria_produto=categoria,
                qtd_sistema=qtd_sistema, qtd_contagem=qtd_contagem, divergencia_qtd=divergencia_qtd,
                valor_estimado=valor_estimado, percentual_acuracia=percentual_acuracia, divergente=divergente,
                resumo_planilha=val(linha, "resumo"), observacao_pos_inventario=val(linha, "obs_pos_inventario"),
                observacao_extra=val(linha, "obs2"), recorrencias_anteriores=recorrencias,
                destaque_recorrente=recorrencias > 0, divergencia_id=divergencia_id,
                movimentacao_historico_id=movimentacao_historico_id,
            )
            db.add(novo_item)
            db.flush()

            obs_pos_inv = val(linha, "obs_pos_inventario")
            if divergente and obs_pos_inv and str(obs_pos_inv).strip():
                db.add(models.AcaoPosInventario(
                    item_fechamento_id=novo_item.id,
                    fechamento_id=fechamento.id, sku=sku, descricao_produto=val(linha, "descricao"),
                    almoxarifado=almoxarifado, acao_descricao=str(obs_pos_inv).strip(),
                    status="Pendente", origem_automatica=True, criado_por=usuario.username,
                ))
            elif not divergente:
                # o item NÃO divergiu neste fechamento - se havia ação
                # pendente de um fechamento anterior pra este mesmo
                # sku+almoxarifado, ela já foi resolvida na prática.
                pendentes_antigas = (
                    db.query(models.AcaoPosInventario)
                    .filter(
                        models.AcaoPosInventario.sku == sku,
                        models.AcaoPosInventario.almoxarifado == almoxarifado,
                        models.AcaoPosInventario.status.in_(["Pendente", "Em_Andamento"]),
                    )
                    .all()
                )
                for antiga in pendentes_antigas:
                    antiga.status = "Concluida"
                    antiga.concluido_em = datetime.utcnow()
                    antiga.observacao_conclusao = (antiga.observacao_conclusao or "") + f" [Fechado automaticamente: sem divergência no fechamento de {data_fechamento}]"
            total += 1
        except Exception as e:
            erros.append(f"linha {i}: {e}")

    fechamento.total_itens = total
    fechamento.total_divergentes = divergentes
    fechamento.valor_total_divergente = round(valor_total, 2)

    registrar_log(db, usuario.username, "importar_fechamento_inventario", entidade="fechamento", entidade_id=fechamento.id,
                  detalhes={"arquivo": arquivo.filename, "total": total, "divergentes": divergentes})
    db.commit()

    return {
        "fechamento_id": fechamento.id,
        "almoxarifado": fechamento.almoxarifado,
        "data_fechamento": str(fechamento.data_fechamento),
        "total_itens": total,
        "total_divergentes": divergentes,
        "valor_total_divergente": fechamento.valor_total_divergente,
        "erros": erros,
    }


@router.get("", response_model=list[schemas.FechamentoOut])
def listar_fechamentos(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    return db.query(models.FechamentoInventario).order_by(models.FechamentoInventario.data_fechamento.desc()).all()


# ==================== Dashboard de acompanhamento ====================

def _query_itens_filtrados(db: Session, almoxarifado: str | None, mes: str | None):
    q = db.query(models.ItemFechamento).join(models.FechamentoInventario, models.ItemFechamento.fechamento_id == models.FechamentoInventario.id)
    if almoxarifado:
        q = q.filter(models.ItemFechamento.almoxarifado == almoxarifado)
    if mes:
        from sqlalchemy import extract
        ano, mes_num = mes.split("-")
        q = q.filter(extract("year", models.FechamentoInventario.data_fechamento) == int(ano))
        q = q.filter(extract("month", models.FechamentoInventario.data_fechamento) == int(mes_num))
    return q


@router.get("/dashboard/kpis")
def dashboard_kpis(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    total = len(itens)
    divergentes = [i for i in itens if i.divergente]
    sem_divergencia = total - len(divergentes)

    acuracia_geral = round(sem_divergencia / total * 100, 2) if total else None
    com_percentual = [i for i in itens if i.percentual_acuracia is not None]
    acima_95 = [i for i in com_percentual if i.percentual_acuracia >= 0.95]
    pct_skus_acima_95 = round(len(acima_95) / len(com_percentual) * 100, 2) if com_percentual else None

    faltas = sum(abs(i.valor_estimado) for i in divergentes if i.divergencia_qtd < 0)
    sobras = sum(abs(i.valor_estimado) for i in divergentes if i.divergencia_qtd > 0)

    return {
        "total_itens": total,
        "total_divergentes": len(divergentes),
        "acuracia_geral_pct": acuracia_geral,
        "pct_skus_acima_95": pct_skus_acima_95,
        "deficit_faltas": round(faltas, 2),
        "total_sobras": round(sobras, 2),
        "resultado_liquido": round(sobras - faltas, 2),
    }


@router.get("/dashboard/por-grupo")
def dashboard_por_grupo(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    from collections import defaultdict
    por_grupo = defaultdict(lambda: {"total": 0, "sem_divergencia": 0, "com_pct": 0, "acima_95": 0})
    for i in itens:
        g = i.categoria_produto or "Sem categoria"
        por_grupo[g]["total"] += 1
        if not i.divergente:
            por_grupo[g]["sem_divergencia"] += 1
        if i.percentual_acuracia is not None:
            por_grupo[g]["com_pct"] += 1
            if i.percentual_acuracia >= 0.95:
                por_grupo[g]["acima_95"] += 1

    resultado = []
    for grupo, v in por_grupo.items():
        resultado.append({
            "grupo": grupo,
            "acuracia_pct": round(v["sem_divergencia"] / v["total"] * 100, 2) if v["total"] else None,
            "pct_skus_acima_95": round(v["acima_95"] / v["com_pct"] * 100, 2) if v["com_pct"] else None,
            "total_itens": v["total"],
        })
    return sorted(resultado, key=lambda x: x["acuracia_pct"] or 0, reverse=True)


@router.get("/dashboard/por-almoxarifado")
def dashboard_por_almoxarifado(mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    itens = _query_itens_filtrados(db, None, mes).all()
    from collections import defaultdict
    por_almox = defaultdict(lambda: {"total": 0, "sem_divergencia": 0})
    for i in itens:
        por_almox[i.almoxarifado]["total"] += 1
        if not i.divergente:
            por_almox[i.almoxarifado]["sem_divergencia"] += 1
    resultado = [
        {"almoxarifado": a, "acuracia_pct": round(v["sem_divergencia"] / v["total"] * 100, 2) if v["total"] else None, "total_itens": v["total"]}
        for a, v in por_almox.items()
    ]
    return sorted(resultado, key=lambda x: x["acuracia_pct"] or 0, reverse=True)


@router.get("/dashboard/ranking-financeiro")
def dashboard_ranking_financeiro(almoxarifado: str | None = None, mes: str | None = None, limite: int = 10, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    itens = _query_itens_filtrados(db, almoxarifado, mes).filter(models.ItemFechamento.divergente.is_(True)).all()
    from collections import defaultdict
    por_sku = defaultdict(lambda: {"descricao": None, "valor": 0.0, "qtd": 0.0})
    for i in itens:
        por_sku[i.sku]["descricao"] = i.descricao_produto
        por_sku[i.sku]["valor"] += i.valor_estimado if i.divergencia_qtd < 0 else -i.valor_estimado
        por_sku[i.sku]["qtd"] += i.divergencia_qtd

    faltas = sorted([(s, v) for s, v in por_sku.items() if v["valor"] > 0], key=lambda x: -x[1]["valor"])[:limite]
    sobras = sorted([(s, v) for s, v in por_sku.items() if v["valor"] < 0], key=lambda x: x[1]["valor"])[:limite]

    return {
        "top_faltas": [{"sku": s, "descricao": v["descricao"], "valor": round(v["valor"], 2)} for s, v in faltas],
        "top_sobras": [{"sku": s, "descricao": v["descricao"], "valor": round(abs(v["valor"]), 2)} for s, v in sobras],
    }


@router.get("/dashboard/top-recorrentes")
def dashboard_top_recorrentes(almoxarifado: str | None = None, limite: int = 10, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Top SKUs que mais vezes apareceram divergentes, somando todos os
    fechamentos já importados - o que você chamaria de 'always suspects'."""
    q = db.query(models.ItemFechamento).filter(models.ItemFechamento.divergente.is_(True))
    if almoxarifado:
        q = q.filter(models.ItemFechamento.almoxarifado == almoxarifado)
    itens = q.all()
    from collections import defaultdict
    por_sku = defaultdict(lambda: {"descricao": None, "ocorrencias": 0, "valor_total": 0.0})
    for i in itens:
        por_sku[i.sku]["descricao"] = i.descricao_produto
        por_sku[i.sku]["ocorrencias"] += 1
        por_sku[i.sku]["valor_total"] += abs(i.valor_estimado)
    ranking = sorted(por_sku.items(), key=lambda x: -x[1]["ocorrencias"])[:limite]
    return [{"sku": s, "descricao": v["descricao"], "ocorrencias": v["ocorrencias"], "valor_total": round(v["valor_total"], 2)} for s, v in ranking]


@router.get("/dashboard/top-impacto-financeiro")
def dashboard_top_impacto_financeiro(almoxarifado: str | None = None, limite: int = 10, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Top SKUs por maior passivo financeiro acumulado - soma só as
    FALTAS (divergencia_qtd < 0, ou seja, perda/déficit real), não
    sobras. Diferente de 'top-recorrentes': aqui o que importa é o
    tamanho do prejuízo, não quantas vezes o item apareceu."""
    q = db.query(models.ItemFechamento).filter(
        models.ItemFechamento.divergente.is_(True), models.ItemFechamento.divergencia_qtd < 0
    )
    if almoxarifado:
        q = q.filter(models.ItemFechamento.almoxarifado == almoxarifado)
    itens = q.all()
    from collections import defaultdict
    por_sku = defaultdict(lambda: {"descricao": None, "ocorrencias": 0, "valor_total": 0.0})
    for i in itens:
        por_sku[i.sku]["descricao"] = i.descricao_produto
        por_sku[i.sku]["ocorrencias"] += 1
        por_sku[i.sku]["valor_total"] += abs(i.valor_estimado)
    ranking = sorted(por_sku.items(), key=lambda x: -x[1]["valor_total"])[:limite]
    return [{"sku": s, "descricao": v["descricao"], "ocorrencias": v["ocorrencias"], "valor_total": round(v["valor_total"], 2)} for s, v in ranking]


@router.get("/dashboard/evolucao-mensal")
def dashboard_evolucao_mensal(almoxarifado: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """MoM: acurácia por mês + quantas fechamentos/almoxarifados foram
    avaliados naquele mês - pra não confundir queda de % com aumento de
    cobertura (mais almoxarifados auditados no período)."""
    q_fech = db.query(models.FechamentoInventario)
    if almoxarifado:
        q_fech = q_fech.filter(models.FechamentoInventario.almoxarifado == almoxarifado)
    fechamentos = q_fech.all()

    from collections import defaultdict
    por_mes = defaultdict(lambda: {"fechamento_ids": set(), "almoxarifados": set()})
    for f in fechamentos:
        mes = f.data_fechamento.strftime("%Y-%m")
        por_mes[mes]["fechamento_ids"].add(f.id)
        por_mes[mes]["almoxarifados"].add(f.almoxarifado)

    resultado = []
    anterior = None
    for mes in sorted(por_mes.keys()):
        ids_fechamento = por_mes[mes]["fechamento_ids"]
        itens = db.query(models.ItemFechamento).filter(models.ItemFechamento.fechamento_id.in_(ids_fechamento)).all()
        total = len(itens)
        sem_div = sum(1 for i in itens if not i.divergente)
        acuracia = round(sem_div / total * 100, 2) if total else None
        valor_divergencia = round(sum(abs(i.valor_estimado) for i in itens if i.divergente), 2)
        variacao_pp = round(acuracia - anterior, 2) if (acuracia is not None and anterior is not None) else None
        resultado.append({
            "mes": mes,
            "acuracia_pct": acuracia,
            "variacao_mom_pp": variacao_pp,
            "qtd_fechamentos_realizados": len(ids_fechamento),
            "qtd_almoxarifados_avaliados": len(por_mes[mes]["almoxarifados"]),
            "valor_divergencia_total": valor_divergencia,
        })
        anterior = acuracia
    return resultado


@router.get("/dashboard/evolucao-por-almox-mensal")
def dashboard_evolucao_por_almox_mensal(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    fechamentos = db.query(models.FechamentoInventario).all()
    from collections import defaultdict
    por_mes_almox = defaultdict(lambda: defaultdict(lambda: {"total": 0, "sem_divergencia": 0}))
    for f in fechamentos:
        mes = f.data_fechamento.strftime("%Y-%m")
        itens = db.query(models.ItemFechamento).filter_by(fechamento_id=f.id).all()
        for i in itens:
            por_mes_almox[mes][f.almoxarifado]["total"] += 1
            if not i.divergente:
                por_mes_almox[mes][f.almoxarifado]["sem_divergencia"] += 1

    resultado = []
    for mes, almoxs in por_mes_almox.items():
        for almox, v in almoxs.items():
            resultado.append({
                "mes": mes, "almoxarifado": almox,
                "acuracia_pct": round(v["sem_divergencia"] / v["total"] * 100, 2) if v["total"] else None,
            })
    return sorted(resultado, key=lambda x: x["mes"])


# ==================== Pós-Inventário: ações de acompanhamento ====================

@router.post("/acoes/confirmar-lote")
def confirmar_acoes_lote(payload: schemas.AcoesLoteAtualizar, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Atualiza várias ações de uma vez - pensado pro caso comum de
    'essas pendências já foram resolvidas há tempos, só ninguém fechou
    elas aqui' - marca todas como Concluída (ou outro status) de uma
    tacada, em vez de uma por uma."""
    if payload.status not in ("Pendente", "Em_Andamento", "Concluida", "Cancelada"):
        raise HTTPException(400, "Status inválido.")
    acoes = db.query(models.AcaoPosInventario).filter(models.AcaoPosInventario.id.in_(payload.ids)).all()
    if not acoes:
        raise HTTPException(404, "Nenhuma ação encontrada com esses IDs.")
    agora = datetime.utcnow()
    for acao in acoes:
        acao.status = payload.status
        if payload.observacao_conclusao:
            acao.observacao_conclusao = payload.observacao_conclusao
        if payload.status == "Concluida" and not acao.concluido_em:
            acao.concluido_em = agora
    registrar_log(db, usuario.username, "confirmar_acoes_lote", detalhes={"quantidade": len(acoes), "novo_status": payload.status})
    db.commit()
    return {"ok": True, "atualizadas": len(acoes)}


@router.post("/acoes/reconciliar-automaticamente")
def reconciliar_acoes_automaticamente(usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Varre as ações Pendente/Em_Andamento e fecha automaticamente as
    que já foram resolvidas na prática - ou seja, o SKU (no mesmo
    almoxarifado) apareceu num fechamento MAIS RECENTE sem divergência.
    Isso é o que corrige o caso de 'pendências' que na verdade já foram
    resolvidas há tempos e ninguém fechou o registro aqui."""
    pendentes = db.query(models.AcaoPosInventario).filter(models.AcaoPosInventario.status.in_(["Pendente", "Em_Andamento"])).all()
    resolvidas = 0
    for acao in pendentes:
        item_origem = db.query(models.ItemFechamento).get(acao.item_fechamento_id) if acao.item_fechamento_id else None
        fechamento_origem = db.query(models.FechamentoInventario).get(acao.fechamento_id) if acao.fechamento_id else None
        data_origem = fechamento_origem.data_fechamento if fechamento_origem else None

        q_posterior = db.query(models.ItemFechamento).join(
            models.FechamentoInventario, models.ItemFechamento.fechamento_id == models.FechamentoInventario.id
        ).filter(
            models.ItemFechamento.sku == acao.sku,
            models.ItemFechamento.almoxarifado == (acao.almoxarifado or (item_origem.almoxarifado if item_origem else None)),
            models.ItemFechamento.divergente.is_(False),
        )
        if data_origem:
            q_posterior = q_posterior.filter(models.FechamentoInventario.data_fechamento > data_origem)
        resolvido_depois = q_posterior.first()

        if resolvido_depois:
            acao.status = "Concluida"
            acao.concluido_em = datetime.utcnow()
            acao.observacao_conclusao = (acao.observacao_conclusao or "") + " [Fechado automaticamente: item não divergiu em fechamento posterior]"
            resolvidas += 1

    registrar_log(db, usuario.username, "reconciliar_acoes_pos_inventario", detalhes={"verificadas": len(pendentes), "resolvidas": resolvidas})
    db.commit()
    return {"verificadas": len(pendentes), "resolvidas_automaticamente": resolvidas}


@router.get("/historico-sku/{sku}")
def historico_sku_fechamento(sku: str, almoxarifado: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Linha do tempo de um SKU através de todos os fechamentos de
    inventário já importados - usado no painel de acompanhamento do
    Pós-Inventário pra mostrar dias movimentados x dias pendentes."""
    q = db.query(models.ItemFechamento).join(
        models.FechamentoInventario, models.ItemFechamento.fechamento_id == models.FechamentoInventario.id
    ).filter(models.ItemFechamento.sku == sku)
    if almoxarifado:
        q = q.filter(models.ItemFechamento.almoxarifado == almoxarifado)
    itens = q.all()

    pontos = []
    for i in itens:
        f = db.query(models.FechamentoInventario).get(i.fechamento_id)
        pontos.append({
            "data": str(f.data_fechamento) if f else None, "almoxarifado": i.almoxarifado,
            "divergente": i.divergente, "resumo": i.resumo_planilha, "observacao_pos_inventario": i.observacao_pos_inventario,
        })
    pontos.sort(key=lambda p: p["data"] or "")

    dias_pendente = sum(1 for p in pontos if p["divergente"])
    dias_resolvido = sum(1 for p in pontos if not p["divergente"])
    return {
        "sku": sku,
        "dias_movimentados": len(pontos),
        "dias_pendente": dias_pendente,
        "dias_resolvido": dias_resolvido,
        "primeira_ocorrencia": pontos[0]["data"] if pontos else None,
        "ultima_ocorrencia": pontos[-1]["data"] if pontos else None,
        "linha_do_tempo": pontos,
    }


@router.post("/recalcular-valores")
def recalcular_valores_fechamento(usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Reaplica o custo unitário cadastrado sobre os itens de fechamento
    divergentes que ainda estão com valor_estimado desatualizado (comum
    quando o custo é importado DEPOIS do fechamento) - mesmo problema e
    mesma solução que já existe pra Divergencia."""
    itens = db.query(models.ItemFechamento).filter(models.ItemFechamento.divergente.is_(True)).all()
    skus = {i.sku for i in itens}
    custos = {p.sku: p.custo_unitario for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus), models.Produto.custo_unitario.isnot(None)).all()}
    atualizados = 0
    for i in itens:
        custo = custos.get(i.sku)
        if custo is None:
            continue
        novo_valor = round(abs(i.divergencia_qtd or 0) * custo, 2)
        if novo_valor != i.valor_estimado:
            i.valor_estimado = novo_valor
            atualizados += 1
    registrar_log(db, usuario.username, "recalcular_valores_fechamento", detalhes={"verificados": len(itens), "atualizados": atualizados})
    db.commit()
    return {"itens_verificados": len(itens), "itens_atualizados": atualizados}


FAIXAS_MAGNITUDE = [(0, 5), (5, 20), (20, 100), (100, float("inf"))]
ROTULOS_MAGNITUDE = ["0 a 5 un.", "5 a 20 un.", "20 a 100 un.", "mais de 100 un."]


@router.get("/dashboard/iap")
def dashboard_iap(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """IAP - Índice de Acurácia Ponderada por valor. Diferente da
    'acurácia geral' (item a item, binária), aqui uma divergência de
    R$ 2.000 pesa muito mais que uma de R$ 5, mesmo que ambas contem
    como '1 item divergente' na métrica clássica. Só entra no cálculo o
    universo de SKUs com custo unitário cadastrado (Cadastros > Produtos,
    ou importado via 'Custos por SKU' / 'Custos (tabela de preço)') -
    cobertura parcial é normal e fica explícita no retorno, não afeta o
    resto do sistema."""
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    if not itens:
        return {"iap_pct": None, "cobertura_custo_pct": None, "itens_avaliados_com_custo": 0, "itens_total": 0, "valor_portfolio_avaliado": 0, "valor_divergente": 0}

    skus = {i.sku for i in itens}
    custos = {
        p.sku: p.custo_unitario
        for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus), models.Produto.custo_unitario.isnot(None)).all()
    }

    itens_com_custo = [i for i in itens if i.sku in custos]
    valor_portfolio = sum((i.qtd_sistema or 0) * custos[i.sku] for i in itens_com_custo)
    # Recalcula o valor divergente na hora, com o MESMO custo do denominador -
    # usar o valor_estimado gravado na importação seria inconsistente (pode
    # ter sido calculado antes do custo existir, ou vir de uma base de custo
    # diferente da planilha original) e quebraria a matemática do índice.
    valor_divergente = sum(abs(i.divergencia_qtd or 0) * custos[i.sku] for i in itens_com_custo if i.divergente)

    iap = round((1 - valor_divergente / valor_portfolio) * 100, 2) if valor_portfolio else None
    return {
        "iap_pct": iap,
        "cobertura_custo_pct": round(len(itens_com_custo) / len(itens) * 100, 2),
        "itens_avaliados_com_custo": len(itens_com_custo),
        "itens_total": len(itens),
        "valor_portfolio_avaliado": round(valor_portfolio, 2),
        "valor_divergente": round(valor_divergente, 2),
    }


@router.get("/dashboard/top-recorrentes-risco")
def dashboard_top_recorrentes_risco(almoxarifado: str | None = None, limite: int = 10, minimo_ocorrencias: int = 2, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Diferente de 'top-recorrentes' (ordenado só por número de
    ocorrências) - aqui o critério é RISCO: precisa ser recorrente
    (mínimo de ocorrências) E o ranking é por ocorrências × valor total
    impactado, não só um dos dois. Um item que aparece 2x com R$ 5.000 de
    impacto sobe mais que um que aparece 8x com R$ 20 de impacto - é isso
    que difere 'recorrência que é só ruído operacional' de 'recorrência
    que é risco real de negócio'."""
    q = db.query(models.ItemFechamento).filter(models.ItemFechamento.divergente.is_(True))
    if almoxarifado:
        q = q.filter(models.ItemFechamento.almoxarifado == almoxarifado)
    itens = q.all()

    from collections import defaultdict
    por_sku = defaultdict(lambda: {"descricao": None, "almoxarifado": None, "ocorrencias": 0, "valor_total": 0.0, "ultima_data": None, "fechamento_id_recente": None, "item_fechamento_id_recente": None})
    for i in itens:
        d = por_sku[i.sku]
        d["descricao"] = i.descricao_produto
        d["almoxarifado"] = i.almoxarifado
        d["ocorrencias"] += 1
        d["valor_total"] += abs(i.valor_estimado or 0)
        f = db.query(models.FechamentoInventario).get(i.fechamento_id)
        data_f = f.data_fechamento if f else None
        if data_f and (d["ultima_data"] is None or data_f > d["ultima_data"]):
            d["ultima_data"] = data_f
            d["fechamento_id_recente"] = i.fechamento_id
            d["item_fechamento_id_recente"] = i.id

    elegiveis = {sku: v for sku, v in por_sku.items() if v["ocorrencias"] >= minimo_ocorrencias}
    for v in elegiveis.values():
        v["score_risco"] = round(v["ocorrencias"] * v["valor_total"], 2)

    ranking = sorted(elegiveis.items(), key=lambda x: -x[1]["score_risco"])[:limite]
    return [
        {
            "sku": s, "descricao": v["descricao"], "almoxarifado": v["almoxarifado"],
            "ocorrencias": v["ocorrencias"], "valor_total": round(v["valor_total"], 2),
            "score_risco": v["score_risco"], "ultima_ocorrencia": str(v["ultima_data"]) if v["ultima_data"] else None,
            "item_fechamento_id_recente": v["item_fechamento_id_recente"],
        }
        for s, v in ranking
    ]


@router.get("/dashboard/iaq")
def dashboard_iaq(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """IAQ - acurácia ponderada por QUANTIDADE. Não depende de custo
    cadastrado (por isso funciona sempre, mesmo com 0% de cobertura de
    custo) - serve de indicador de transição enquanto o cadastro de custo
    não está completo, e de checagem cruzada do IAP quando já está."""
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    if not itens:
        return {"iaq_pct": None, "qtd_sistema_total": 0, "qtd_divergente_total": 0}
    qtd_sistema_total = sum(abs(i.qtd_sistema or 0) for i in itens)
    qtd_divergente_total = sum(abs(i.divergencia_qtd or 0) for i in itens if i.divergente)
    iaq = round((1 - qtd_divergente_total / qtd_sistema_total) * 100, 2) if qtd_sistema_total else None
    return {"iaq_pct": iaq, "qtd_sistema_total": round(qtd_sistema_total, 2), "qtd_divergente_total": round(qtd_divergente_total, 2)}


@router.get("/dashboard/comparativo-acuracia")
def dashboard_comparativo_acuracia(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """O número que prova a distorção: os três modelos lado a lado, sobre
    o MESMO universo de itens. O gap entre item-a-item e os ponderados é
    o tamanho real da distorção que a métrica clássica escondia."""
    kpis = dashboard_kpis(almoxarifado, mes, usuario, db)
    iaq = dashboard_iaq(almoxarifado, mes, usuario, db)
    iap = dashboard_iap(almoxarifado, mes, usuario, db)

    item_a_item = kpis["acuracia_geral_pct"]
    resultado = {
        "item_a_item_pct": item_a_item,
        "iaq_pct": iaq["iaq_pct"],
        "iap_pct": iap["iap_pct"],
        "cobertura_custo_pct": iap["cobertura_custo_pct"],
        "gap_item_vs_iaq_pp": round(iaq["iaq_pct"] - item_a_item, 2) if (iaq["iaq_pct"] is not None and item_a_item is not None) else None,
        "gap_item_vs_iap_pp": round(iap["iap_pct"] - item_a_item, 2) if (iap["iap_pct"] is not None and item_a_item is not None) else None,
    }
    return resultado


@router.get("/dashboard/concentracao-valor")
def dashboard_concentracao_valor(almoxarifado: str | None = None, mes: str | None = None, top_n: int = 10, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Curva de Pareto: quanto do valor total em risco está concentrado
    nos itens de maior impacto. Isso é a prova visual de que 'tratar
    tudo igual' (modelo item a item) ignora que poucos SKUs concentram
    quase todo o risco financeiro real."""
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    divergentes = [i for i in itens if i.divergente]
    if not divergentes:
        return {"itens": [], "top_n_pct_do_valor": None, "top_n": top_n, "total_itens_divergentes": 0}

    skus = {i.sku for i in divergentes}
    custos = {p.sku: p.custo_unitario for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus), models.Produto.custo_unitario.isnot(None)).all()}

    def valor_atual(item):
        custo = custos.get(item.sku)
        if custo is not None:
            return abs(item.divergencia_qtd or 0) * custo
        return abs(item.valor_estimado or 0)  # fallback: usa o valor gravado (ex: vindo direto do "Vl" da planilha) se não há custo cadastrado

    ordenados = sorted(divergentes, key=valor_atual, reverse=True)
    valor_total = sum(valor_atual(i) for i in ordenados)

    curva = []
    acumulado = 0.0
    for idx, item in enumerate(ordenados, start=1):
        acumulado += valor_atual(item)
        curva.append({
            "posicao": idx, "sku": item.sku, "descricao": item.descricao_produto, "almoxarifado": item.almoxarifado,
            "valor": round(valor_atual(item), 2),
            "pct_itens_acumulado": round(idx / len(ordenados) * 100, 1),
            "pct_valor_acumulado": round(acumulado / valor_total * 100, 1) if valor_total else 0,
        })

    top_n_valor = sum(valor_atual(i) for i in ordenados[:top_n])
    return {
        "itens": curva[:50],  # a curva completa raramente precisa de mais que isso pro grafico
        "total_itens_divergentes": len(ordenados),
        "valor_total": round(valor_total, 2),
        "top_n": min(top_n, len(ordenados)),
        "top_n_pct_do_valor": round(top_n_valor / valor_total * 100, 1) if valor_total else None,
    }


@router.get("/dashboard/distribuicao-magnitude")
def dashboard_distribuicao_magnitude(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Quantas divergências são 'pequenas' (pouco impacto real) vs
    'grandes' - isso é a prova visual da segunda distorção: a métrica
    item a item trata as duas categorias como se fossem a mesma coisa."""
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    divergentes = [i for i in itens if i.divergente]

    skus = {i.sku for i in divergentes}
    custos = {p.sku: p.custo_unitario for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus), models.Produto.custo_unitario.isnot(None)).all()}

    def valor_atual(item):
        custo = custos.get(item.sku)
        if custo is not None:
            return abs(item.divergencia_qtd or 0) * custo
        return abs(item.valor_estimado or 0)

    faixas = FAIXAS_MAGNITUDE
    rotulos = ROTULOS_MAGNITUDE
    resultado = []
    for (minimo, maximo), rotulo in zip(faixas, rotulos):
        do_grupo = [i for i in divergentes if minimo <= abs(i.divergencia_qtd or 0) < maximo]
        resultado.append({
            "faixa": rotulo, "quantidade_itens": len(do_grupo),
            "valor_total": round(sum(valor_atual(i) for i in do_grupo), 2),
        })

    total = len(divergentes)
    pequenas = resultado[0]["quantidade_itens"] if resultado else 0
    return {
        "faixas": resultado,
        "total_divergentes": total,
        "pct_divergencias_pequenas": round(pequenas / total * 100, 1) if total else None,
    }


@router.get("/dashboard/itens-por-magnitude")
def dashboard_itens_por_magnitude(faixa_idx: int, almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Lista os itens de uma faixa específica do gráfico de distribuição
    por magnitude - alimenta o duplo clique na coluna, pra ver quem são
    os produtos daquela faixa e, se fizer sentido, criar ação de
    acompanhamento pra eles."""
    if faixa_idx < 0 or faixa_idx >= len(FAIXAS_MAGNITUDE):
        raise HTTPException(400, f"faixa_idx deve estar entre 0 e {len(FAIXAS_MAGNITUDE) - 1}")
    minimo, maximo = FAIXAS_MAGNITUDE[faixa_idx]

    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    divergentes = [i for i in itens if i.divergente and minimo <= abs(i.divergencia_qtd or 0) < maximo]

    skus = {i.sku for i in divergentes}
    custos = {p.sku: p.custo_unitario for p in db.query(models.Produto).filter(models.Produto.sku.in_(skus), models.Produto.custo_unitario.isnot(None)).all()}

    def valor_atual(item):
        custo = custos.get(item.sku)
        if custo is not None:
            return abs(item.divergencia_qtd or 0) * custo
        return abs(item.valor_estimado or 0)

    itens_ordenados = sorted(divergentes, key=valor_atual, reverse=True)
    return {
        "faixa": ROTULOS_MAGNITUDE[faixa_idx],
        "itens": [
            {
                "sku": i.sku, "descricao": i.descricao_produto, "almoxarifado": i.almoxarifado,
                "divergencia_qtd": i.divergencia_qtd, "valor": round(valor_atual(i), 2),
            }
            for i in itens_ordenados
        ],
    }


def _tres_modelos(itens: list) -> dict:
    """Calcula os 3 indicadores (item a item, IAQ, IAP) sobre uma lista
    de ItemFechamento já filtrada - reaproveitado por grupo, almoxarifado
    e evolução mensal, pra não repetir a lógica de cálculo em cada lugar."""
    total = len(itens)
    if not total:
        return {"item_a_item_pct": None, "iaq_pct": None, "iap_pct": None}

    divergentes = [i for i in itens if i.divergente]
    sem_divergencia = total - len(divergentes)
    item_a_item = round(sem_divergencia / total * 100, 2)

    qtd_sistema_total = sum(abs(i.qtd_sistema or 0) for i in itens)
    qtd_divergente_total = sum(abs(i.divergencia_qtd or 0) for i in divergentes)
    iaq = round((1 - qtd_divergente_total / qtd_sistema_total) * 100, 2) if qtd_sistema_total else None

    skus = {i.sku for i in itens}
    custos = {}
    if skus:
        from ..database import SessionLocal
        db_tmp = SessionLocal()
        try:
            custos = {p.sku: p.custo_unitario for p in db_tmp.query(models.Produto).filter(models.Produto.sku.in_(skus), models.Produto.custo_unitario.isnot(None)).all()}
        finally:
            db_tmp.close()
    itens_com_custo = [i for i in itens if i.sku in custos]
    if itens_com_custo:
        valor_portfolio = sum((i.qtd_sistema or 0) * custos[i.sku] for i in itens_com_custo)
        valor_divergente = sum(abs(i.divergencia_qtd or 0) * custos[i.sku] for i in itens_com_custo if i.divergente)
        iap = round((1 - valor_divergente / valor_portfolio) * 100, 2) if valor_portfolio else None
    else:
        iap = None

    return {"item_a_item_pct": item_a_item, "iaq_pct": iaq, "iap_pct": iap}


@router.get("/dashboard/comparativo-por-grupo")
def dashboard_comparativo_por_grupo(almoxarifado: str | None = None, mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Os 3 modelos (item a item, IAQ, IAP) lado a lado, por categoria de
    produto - pra ver em qual grupo a distorção é maior."""
    itens = _query_itens_filtrados(db, almoxarifado, mes).all()
    from collections import defaultdict
    por_grupo = defaultdict(list)
    for i in itens:
        por_grupo[i.categoria_produto or "Sem categoria"].append(i)

    resultado = [{"grupo": grupo, **_tres_modelos(lista)} for grupo, lista in por_grupo.items()]
    return sorted(resultado, key=lambda x: x["item_a_item_pct"] or 0, reverse=True)


@router.get("/dashboard/comparativo-por-almoxarifado")
def dashboard_comparativo_por_almoxarifado(mes: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Os 3 modelos lado a lado, por almoxarifado."""
    itens = _query_itens_filtrados(db, None, mes).all()
    from collections import defaultdict
    por_almox = defaultdict(list)
    for i in itens:
        por_almox[i.almoxarifado].append(i)

    resultado = [{"almoxarifado": almox, **_tres_modelos(lista)} for almox, lista in por_almox.items()]
    return sorted(resultado, key=lambda x: x["item_a_item_pct"] or 0, reverse=True)


@router.get("/dashboard/evolucao-ponderada-mensal")
def dashboard_evolucao_ponderada_mensal(almoxarifado: str | None = None, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """IAQ e IAP mês a mês, com variação MoM em pontos percentuais - pra
    ver se a distância entre os modelos está diminuindo (bom sinal: itens
    de maior impacto sendo resolvidos primeiro) ou aumentando (alerta)."""
    fechamentos = db.query(models.FechamentoInventario).all()
    if almoxarifado:
        fechamentos = [f for f in fechamentos if f.almoxarifado == almoxarifado]

    meses = sorted({f.data_fechamento.strftime("%Y-%m") for f in fechamentos})
    resultado = []
    anterior = {"item_a_item_pct": None, "iaq_pct": None, "iap_pct": None}
    for mes in meses:
        iaq = dashboard_iaq(almoxarifado, mes, usuario, db)
        iap = dashboard_iap(almoxarifado, mes, usuario, db)
        kpis = dashboard_kpis(almoxarifado, mes, usuario, db)
        atual = {"item_a_item_pct": kpis["acuracia_geral_pct"], "iaq_pct": iaq["iaq_pct"], "iap_pct": iap["iap_pct"]}

        def variacao(chave):
            if atual[chave] is None or anterior[chave] is None:
                return None
            return round(atual[chave] - anterior[chave], 2)

        resultado.append({
            "mes": mes, **atual,
            "variacao_item_pp": variacao("item_a_item_pct"),
            "variacao_iaq_pp": variacao("iaq_pct"),
            "variacao_iap_pp": variacao("iap_pct"),
        })
        anterior = atual
    return resultado


@router.get("/acoes", response_model=list[schemas.AcaoPosInventarioOut])
def listar_acoes(
    status: str | None = None,
    responsavel: str | None = None,
    fechamento_id: int | None = None,
    sku: str | None = None,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    q = db.query(models.AcaoPosInventario)
    if status:
        q = q.filter_by(status=status)
    if responsavel:
        q = q.filter_by(responsavel=responsavel)
    if fechamento_id:
        q = q.filter_by(fechamento_id=fechamento_id)
    if sku:
        q = q.filter_by(sku=sku)
    return q.order_by(models.AcaoPosInventario.criado_em.desc()).all()


@router.get("/acoes/{acao_id}", response_model=schemas.AcaoPosInventarioOut)
def detalhar_acao(acao_id: int, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    acao = db.query(models.AcaoPosInventario).get(acao_id)
    if not acao:
        raise HTTPException(404, "Ação não encontrada.")
    return acao


@router.post("/acoes", response_model=schemas.AcaoPosInventarioOut)
def criar_acao(payload: schemas.AcaoPosInventarioCreate, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    nova = models.AcaoPosInventario(**payload.model_dump(), status="Pendente", criado_por=usuario.username)
    db.add(nova)
    registrar_log(db, usuario.username, "criar_acao_pos_inventario", entidade="acao_pos_inventario", entidade_id=payload.sku)
    db.commit()
    db.refresh(nova)
    return nova


@router.patch("/acoes/{acao_id}", response_model=schemas.AcaoPosInventarioOut)
def atualizar_acao(acao_id: int, payload: schemas.AcaoPosInventarioAtualizar, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    acao = db.query(models.AcaoPosInventario).get(acao_id)
    if not acao:
        raise HTTPException(404, "Ação não encontrada.")
    dados = payload.model_dump(exclude_unset=True)
    for campo, valor in dados.items():
        setattr(acao, campo, valor)
    if dados.get("status") == "Concluida" and not acao.concluido_em:
        acao.concluido_em = datetime.utcnow()
    registrar_log(db, usuario.username, "atualizar_acao_pos_inventario", entidade="acao_pos_inventario", entidade_id=acao_id, detalhes=dados)
    db.commit()
    db.refresh(acao)
    return acao


@router.delete("/acoes/{acao_id}")
def excluir_acao(acao_id: int, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    acao = db.query(models.AcaoPosInventario).get(acao_id)
    if not acao:
        raise HTTPException(404, "Ação não encontrada.")
    registrar_log(db, usuario.username, "excluir_acao_pos_inventario", entidade="acao_pos_inventario", entidade_id=acao_id)
    db.delete(acao)
    db.commit()
    return {"ok": True}


# ==================== Rotas dinâmicas de fechamento (registradas por ==================
# último de propósito: "/{fechamento_id}" é um padrão de 1 segmento que
# colidiria com rotas literais como "/acoes" se viesse antes - o roteador
# do FastAPI casa por ordem de registro, não por especificidade. ==========

@router.get("/ciencia/{ciencia_id}/pdf")
def baixar_pdf_ciencia(ciencia_id: int, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    from fastapi.responses import Response
    from ..pdf_ciencia import gerar_pdf_ciencia

    ciencia = db.query(models.ConciliacaoCiencia).get(ciencia_id)
    if not ciencia:
        raise HTTPException(404, "Registro de ciência não encontrado.")
    fechamento = db.query(models.FechamentoInventario).get(ciencia.fechamento_id)
    if not fechamento:
        raise HTTPException(404, "Fechamento de origem não encontrado.")

    pdf_bytes = gerar_pdf_ciencia(ciencia, fechamento)
    nome_arquivo = f"ciencia_fechamento_{fechamento.id}_{ciencia.data_assinatura.strftime('%Y%m%d')}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{nome_arquivo}"'})


@router.get("/{fechamento_id}", response_model=schemas.FechamentoOut)
def detalhar_fechamento(fechamento_id: int, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    f = db.query(models.FechamentoInventario).get(fechamento_id)
    if not f:
        raise HTTPException(404, "Fechamento não encontrado.")
    return f


@router.get("/{fechamento_id}/itens", response_model=list[schemas.ItemFechamentoOut])
def listar_itens_fechamento(
    fechamento_id: int,
    divergente: bool | None = None,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Lista os itens de um fechamento. Passe ?divergente=true pra pegar
    só a lista de divergências (separada dos itens ok), como pedido: ao
    abrir o caso pra investigar, a tela já mostra essa lista isolada."""
    q = db.query(models.ItemFechamento).filter_by(fechamento_id=fechamento_id)
    if divergente is not None:
        q = q.filter_by(divergente=divergente)
    return q.order_by(models.ItemFechamento.destaque_recorrente.desc(), models.ItemFechamento.valor_estimado.desc()).all()


@router.post("/{fechamento_id}/ciencia", response_model=schemas.ConciliacaoCienciaOut)
def gerar_ciencia(fechamento_id: int, payload: schemas.ConciliacaoCienciaCreate, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Confirmação autenticada de um gestor de que revisou a conciliação
    - a 'assinatura' aqui é o usuário logado + timestamp, não uma
    assinatura manuscrita. Congela os itens divergentes ATUAIS num
    snapshot, pra o documento gerado nunca mudar depois."""
    fechamento = db.query(models.FechamentoInventario).get(fechamento_id)
    if not fechamento:
        raise HTTPException(404, "Fechamento não encontrado.")

    itens_divergentes = db.query(models.ItemFechamento).filter_by(fechamento_id=fechamento_id, divergente=True).all()
    snapshot = [
        {
            "sku": i.sku, "descricao": i.descricao_produto, "almoxarifado": i.almoxarifado,
            "qtd_sistema": i.qtd_sistema, "qtd_contagem": i.qtd_contagem,
            "divergencia_qtd": i.divergencia_qtd, "valor_estimado": i.valor_estimado,
        }
        for i in itens_divergentes
    ]
    valor_total = sum(abs(i.valor_estimado or 0) for i in itens_divergentes)

    ciencia = models.ConciliacaoCiencia(
        fechamento_id=fechamento_id, gestor_username=usuario.username, gestor_nome=usuario.nome_exibicao,
        observacao=payload.observacao, itens_divergentes_snapshot=snapshot,
        total_itens_divergentes=len(itens_divergentes), valor_total_divergente=round(valor_total, 2),
    )
    db.add(ciencia)
    registrar_log(db, usuario.username, "gerar_ciencia_conciliacao", entidade="fechamento", entidade_id=fechamento_id,
                  detalhes={"total_itens_divergentes": len(itens_divergentes), "valor_total": round(valor_total, 2)})
    db.commit()
    db.refresh(ciencia)
    return ciencia


@router.get("/{fechamento_id}/ciencia", response_model=list[schemas.ConciliacaoCienciaOut])
def listar_ciencia(fechamento_id: int, usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    return db.query(models.ConciliacaoCiencia).filter_by(fechamento_id=fechamento_id).order_by(models.ConciliacaoCiencia.data_assinatura.desc()).all()


@router.patch("/{fechamento_id}/corrigir-almoxarifado")
def corrigir_almoxarifado_fechamento(
    fechamento_id: int,
    novo_almoxarifado: str,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Corrige o almoxarifado de um fechamento já importado (útil quando
    a planilha tinha um texto de almoxarifado ambíguo que foi
    interpretado errado - ex: um sub-almoxarifado "Processo" que também
    menciona "Fabrica" no nome do site). Propaga a correção para todos
    os itens do fechamento e para as Divergencias/MovimentacoesHistorico
    ligadas a eles, então nada fica com o almoxarifado antigo por engano."""
    fechamento = db.query(models.FechamentoInventario).get(fechamento_id)
    if not fechamento:
        raise HTTPException(404, "Fechamento não encontrado.")
    if not db.query(models.Almoxarifado).filter_by(codigo=novo_almoxarifado).first():
        raise HTTPException(400, f"Almoxarifado '{novo_almoxarifado}' não existe no cadastro.")

    almoxarifado_antigo = fechamento.almoxarifado
    fechamento.almoxarifado = novo_almoxarifado

    itens = db.query(models.ItemFechamento).filter_by(fechamento_id=fechamento_id).all()
    ids_divergencia = [i.divergencia_id for i in itens if i.divergencia_id]
    ids_historico = [i.movimentacao_historico_id for i in itens if i.movimentacao_historico_id]

    for item in itens:
        item.almoxarifado = novo_almoxarifado
    if ids_divergencia:
        db.query(models.Divergencia).filter(models.Divergencia.id.in_(ids_divergencia)).update(
            {"almoxarifado": novo_almoxarifado}, synchronize_session=False
        )
    if ids_historico:
        db.query(models.MovimentacaoHistorico).filter(models.MovimentacaoHistorico.id.in_(ids_historico)).update(
            {"almoxarifado": novo_almoxarifado}, synchronize_session=False
        )

    registrar_log(db, usuario.username, "corrigir_almoxarifado_fechamento", entidade="fechamento", entidade_id=fechamento_id,
                  detalhes={"de": almoxarifado_antigo, "para": novo_almoxarifado, "itens_afetados": len(itens)})
    db.commit()
    return {"ok": True, "itens_corrigidos": len(itens), "almoxarifado_anterior": almoxarifado_antigo, "almoxarifado_novo": novo_almoxarifado}


@router.delete("/{fechamento_id}")
def excluir_fechamento(fechamento_id: int, usuario: models.Usuario = Depends(requer_papel("admin", "analista")), db: Session = Depends(get_db)):
    """Desfaz um fechamento inteiro: remove os itens, as divergências e o
    histórico ligados a ele, as ações pós-inventário associadas, e o
    feedback de ML das divergências removidas."""
    fechamento = db.query(models.FechamentoInventario).get(fechamento_id)
    if not fechamento:
        raise HTTPException(404, "Fechamento não encontrado.")

    itens = db.query(models.ItemFechamento).filter_by(fechamento_id=fechamento_id).all()
    ids_divergencias = [i.divergencia_id for i in itens if i.divergencia_id]
    ids_historico = [i.movimentacao_historico_id for i in itens if i.movimentacao_historico_id]

    if ids_divergencias:
        db.query(models.CasoMLFeedback).filter(models.CasoMLFeedback.divergencia_id.in_(ids_divergencias)).delete(synchronize_session=False)
        db.query(models.Divergencia).filter(models.Divergencia.id.in_(ids_divergencias)).delete(synchronize_session=False)
    if ids_historico:
        db.query(models.MovimentacaoHistorico).filter(models.MovimentacaoHistorico.id.in_(ids_historico)).delete(synchronize_session=False)

    qtd_acoes = db.query(models.AcaoPosInventario).filter_by(fechamento_id=fechamento_id).delete(synchronize_session=False)
    qtd_itens = db.query(models.ItemFechamento).filter_by(fechamento_id=fechamento_id).delete(synchronize_session=False)

    registrar_log(db, usuario.username, "excluir_fechamento_inventario", entidade="fechamento", entidade_id=fechamento_id,
                  detalhes={"arquivo": fechamento.arquivo_origem, "itens_removidos": qtd_itens, "acoes_removidas": qtd_acoes})
    db.delete(fechamento)
    db.commit()
    return {"ok": True, "itens_removidos": qtd_itens, "divergencias_removidas": len(ids_divergencias), "acoes_removidas": qtd_acoes}
