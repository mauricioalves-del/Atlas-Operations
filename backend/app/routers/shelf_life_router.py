import io
from datetime import date, datetime
from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
import openpyxl

from .. import models
from ..database import get_db
from ..csv_utils import parse_sku, parse_decimal, limpar_texto
from ..hipoteses_config import normalizar_almoxarifado
from .. import shelf_life
from ..deps import requer_papel, obter_usuario_atual
from ..audit import registrar_log

router = APIRouter(prefix="/shelf-life", tags=["shelf_life"])


class LoteManualIn(BaseModel):
    sku: str
    descricao_produto: str | None = None
    tipo_material: str | None = None
    almoxarifado: str | None = None  # aceita já normalizado (Almox_...) ou bruto (normaliza aqui)
    lote: str | None = None
    quantidade: float
    unidade: str | None = None
    data_validade: date | None = None
    peso_kg: float | None = None
    custo_unitario: float | None = None
    ativo: bool = True


@router.get("/resumo")
def resumo_shelf_life(usuario: models.Usuario = Depends(obter_usuario_atual), db: Session = Depends(get_db)):
    """Farol de risco de validade (vencido / 30 / 60 / 90 dias / sem
    validade cadastrada) - alimenta a tela dedicada Shelf Life. O mesmo
    cálculo é reaproveitado pelo Mapa de Demandas (tela Início) chamando
    shelf_life.calcular_resumo_shelf_life diretamente, sem passar por
    HTTP - ver dashboard_router.mapa_demandas."""
    return shelf_life.calcular_resumo_shelf_life(db)


@router.get("/lotes")
def listar_lotes_shelf_life(
    almoxarifado: str | None = None,
    farol: str | None = None,
    busca: str | None = None,
    somente_ativos: bool = True,
    usuario: models.Usuario = Depends(obter_usuario_atual),
    db: Session = Depends(get_db),
):
    """Lista lotes com filtros - farol é calculado em memória (não é uma
    coluna do banco, ver shelf_life.calcular_farol) porque precisa sempre
    refletir 'dias até vencer a partir de hoje', não um valor congelado no
    momento da importação."""
    q = db.query(models.LoteShelfLife)
    if somente_ativos:
        q = q.filter(models.LoteShelfLife.ativo.is_(True))
    if almoxarifado:
        q = q.filter(models.LoteShelfLife.almoxarifado == almoxarifado)
    if busca:
        termo = f"%{busca.strip()}%"
        q = q.filter((models.LoteShelfLife.sku.ilike(termo)) | (models.LoteShelfLife.descricao_produto.ilike(termo)) | (models.LoteShelfLife.lote.ilike(termo)))

    hoje = date.today()
    resultado = []
    for l in q.all():
        f = shelf_life.calcular_farol(l.data_validade, hoje)
        if farol and f != farol:
            continue
        dias = (l.data_validade - hoje).days if l.data_validade else None
        resultado.append({
            "id": l.id, "sku": l.sku, "descricao_produto": l.descricao_produto,
            "tipo_material": l.tipo_material, "almoxarifado": l.almoxarifado,
            "almoxarifado_origem": l.almoxarifado_origem, "lote": l.lote,
            "quantidade": l.quantidade, "unidade": l.unidade,
            "data_validade": str(l.data_validade) if l.data_validade else None,
            "dias_para_vencer": dias,
            "peso_kg": l.peso_kg, "custo_unitario": l.custo_unitario,
            "valor_estimado": round((l.quantidade or 0) * (l.custo_unitario or 0), 2),
            "ativo": l.ativo, "origem_cadastro": l.origem_cadastro, "farol": f,
        })
    # sem validade cadastrada vai pro final da lista, não pro topo (senão
    # apareceriam "antes" dos vencidos numa ordenação numérica ingênua)
    resultado.sort(key=lambda x: x["dias_para_vencer"] if x["dias_para_vencer"] is not None else 999999)
    return resultado


@router.post("/lotes")
def criar_lote_manual(
    payload: LoteManualIn,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Cadastro manual de um lote - complementa a importação da planilha
    (Lote_Sistema.xlsx) pra lotes recebidos entre uma exportação e outra,
    ou pra corrigir uma validade errada sem depender de reimportar tudo."""
    sku = parse_sku(payload.sku)
    if not sku:
        raise HTTPException(400, "SKU é obrigatório.")
    almoxarifado_normalizado = None
    if payload.almoxarifado:
        # já vem normalizado (Almox_... ou NAO_MAPEADO__...) se veio de um
        # <select> preenchido a partir do cadastro de almoxarifados; se
        # vier de texto livre, normaliza aqui do mesmo jeito que a
        # importação de planilha faz.
        v = payload.almoxarifado.strip()
        almoxarifado_normalizado = v if (v.startswith("Almox_") or v.startswith("NAO_MAPEADO__")) else normalizar_almoxarifado(v)

    lote = models.LoteShelfLife(
        sku=sku, descricao_produto=limpar_texto(payload.descricao_produto),
        tipo_material=limpar_texto(payload.tipo_material),
        almoxarifado=almoxarifado_normalizado, almoxarifado_origem=payload.almoxarifado,
        lote=limpar_texto(payload.lote), quantidade=payload.quantidade,
        unidade=limpar_texto(payload.unidade), data_validade=payload.data_validade,
        peso_kg=payload.peso_kg, custo_unitario=payload.custo_unitario,
        ativo=payload.ativo, origem_cadastro="manual", criado_por=usuario.username,
    )
    db.add(lote)
    db.flush()
    registrar_log(db, usuario.username, "criar_lote_shelf_life", entidade="lote_shelf_life", entidade_id=lote.id,
                  detalhes={"sku": sku, "lote": lote.lote, "almoxarifado": almoxarifado_normalizado})
    db.commit()
    return {"ok": True, "id": lote.id}


@router.patch("/lotes/{lote_id}")
def editar_lote_shelf_life(
    lote_id: int,
    payload: LoteManualIn,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Correção pontual de um lote (ex: validade errada, quantidade
    desatualizada) - independente de ele ter vindo de importação ou
    cadastro manual."""
    lote = db.query(models.LoteShelfLife).get(lote_id)
    if not lote:
        raise HTTPException(404, "Lote não encontrado.")
    almoxarifado_normalizado = lote.almoxarifado
    if payload.almoxarifado:
        v = payload.almoxarifado.strip()
        almoxarifado_normalizado = v if (v.startswith("Almox_") or v.startswith("NAO_MAPEADO__")) else normalizar_almoxarifado(v)

    lote.sku = parse_sku(payload.sku) or lote.sku
    lote.descricao_produto = limpar_texto(payload.descricao_produto)
    lote.tipo_material = limpar_texto(payload.tipo_material)
    lote.almoxarifado = almoxarifado_normalizado
    lote.lote = limpar_texto(payload.lote)
    lote.quantidade = payload.quantidade
    lote.unidade = limpar_texto(payload.unidade)
    lote.data_validade = payload.data_validade
    lote.peso_kg = payload.peso_kg
    lote.custo_unitario = payload.custo_unitario
    lote.ativo = payload.ativo
    lote.atualizado_em = datetime.utcnow()

    registrar_log(db, usuario.username, "editar_lote_shelf_life", entidade="lote_shelf_life", entidade_id=lote.id)
    db.commit()
    return {"ok": True}


@router.delete("/lotes/{lote_id}")
def excluir_lote_shelf_life(
    lote_id: int,
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    lote = db.query(models.LoteShelfLife).get(lote_id)
    if not lote:
        raise HTTPException(404, "Lote não encontrado.")
    registrar_log(db, usuario.username, "excluir_lote_shelf_life", entidade="lote_shelf_life", entidade_id=lote.id,
                  detalhes={"sku": lote.sku, "lote": lote.lote})
    db.delete(lote)
    db.commit()
    return {"ok": True}


@router.post("/importar-planilha")
async def importar_planilha_lote_sistema(
    arquivo: UploadFile = File(...),
    aba: str | None = Form("Lote_Sistema"),
    usuario: models.Usuario = Depends(requer_papel("admin", "analista")),
    db: Session = Depends(get_db),
):
    """Importa a aba 'Lote_Sistema' do arquivo exportado do sistema
    interno (ex: Lote_Sistema.xlsx) - upsert por (sku, lote, almoxarifado),
    ver shelf_life.importar_linhas_lote_sistema. Não apaga lotes
    existentes que não estão nesta planilha (não é substituição
    completa - ver comentário na função de upsert)."""
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Não consegui abrir o arquivo Excel: {e}")

    nome_aba = aba or wb.sheetnames[0]
    if nome_aba not in wb.sheetnames:
        raise HTTPException(400, f"Aba '{nome_aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
    ws = wb[nome_aba]
    linhas_brutas = list(ws.iter_rows(values_only=True))
    if not linhas_brutas:
        raise HTTPException(400, "Planilha vazia.")
    cabecalho = [str(c).strip() if c is not None else "" for c in linhas_brutas[0]]
    faltando = [c for c in shelf_life.COLUNAS_OBRIGATORIAS_LOTE_SISTEMA if c not in cabecalho]
    if faltando:
        raise HTTPException(400, f"Colunas obrigatórias não encontradas: {faltando}. Cabeçalho encontrado: {cabecalho}")

    linhas = [dict(zip(cabecalho, linha)) for linha in linhas_brutas[1:] if any(v is not None for v in linha)]
    resultado = shelf_life.importar_linhas_lote_sistema(db, linhas, usuario.username)

    registrar_log(db, usuario.username, "importar_shelf_life_planilha",
                  detalhes={"arquivo": arquivo.filename, "aba": nome_aba, **resultado})
    db.commit()
    return {"arquivo": arquivo.filename, "aba": nome_aba, "linhas_na_planilha": len(linhas), **resultado}
